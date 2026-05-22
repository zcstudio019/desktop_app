from __future__ import annotations

import re
import logging
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from .normalizer import normalize_account_number, normalize_amount, normalize_currency, normalize_date, normalize_text


logger = logging.getLogger(__name__)

COUNTERPARTY_ACCOUNT_HEADERS = ("对方账号", "对方账户", "收付款方账号", "交易对手账号", "对手方账号")
COUNTERPARTY_BANK_HEADERS = ("对方开户行", "对方机构", "对方银行", "收付款方开户行", "对方账户开户行", "counterparty_bank")
OWN_ACCOUNT_HEADERS = ("账号", "本方账号", "企业账号", "账户账号", "银行账号", "账户号码", "account_number")

HEADER_SYNONYMS: dict[str, tuple[str, ...]] = {
    "transaction_date": ("交易日期", "记账日期", "入账日期", "交易时间", "发生日期", "日期", "提交时间", "transaction_date", "date"),
    "post_date": ("记账日期", "入账日期", "过账日期", "清算日期", "post_date", "posting_date"),
    "credit_amount": ("贷方金额", "收入", "转入", "入账金额", "发生额收入", "收方金额", "来账金额", "贷方发生额", "收入金额", "借方金额（收）", "借方金额(收)", "credit", "credit_amount", "inflow"),
    "debit_amount": ("借方金额", "支出", "转出", "出账金额", "发生额支出", "付方金额", "往账金额", "借方发生额", "支出金额", "贷方金额（支）", "贷方金额(支)", "debit", "debit_amount", "outflow"),
    "balance": ("余额", "账户余额", "交易后余额", "可用余额", "balance"),
    "summary": ("摘要", "交易摘要", "交易名称", "备注", "附言", "客户附言", "summary", "remark"),
    "purpose": ("用途", "交易用途", "附言", "备注", "purpose", "usage"),
    "counterparty_name": ("对方户名", "对方名称", "交易对手", "对方账户名称", "收付款方名称", "counterparty_name", "counterparty"),
    "counterparty_account": COUNTERPARTY_ACCOUNT_HEADERS,
    "counterparty_bank": COUNTERPARTY_BANK_HEADERS,
    "account_number": OWN_ACCOUNT_HEADERS,
    "currency": ("币种", "货币", "currency"),
}

SHEET_BANK_ALIASES: tuple[tuple[str, str], ...] = (
    ("民生", "民生银行"),
    ("平安", "平安银行"),
    ("泰隆", "泰隆银行"),
    ("浙江网商", "浙江网商"),
    ("网商", "浙江网商"),
)


def infer_bank_from_sheet_name(sheet_name: Any) -> str | None:
    text = normalize_text(sheet_name)
    for keyword, bank_name in SHEET_BANK_ALIASES:
        if keyword in text:
            return bank_name
    if "银行" in text:
        return text
    return None


def _compact(value: Any) -> str:
    return normalize_text(value).replace(" ", "")


def _canonical_header(value: Any) -> str | None:
    compact = _compact(value)
    if not compact:
        return None
    if any(name in compact for name in COUNTERPARTY_BANK_HEADERS):
        return "counterparty_bank"
    if any(name in compact for name in COUNTERPARTY_ACCOUNT_HEADERS):
        return "counterparty_account"
    if "借方金额" in compact and "收" in compact:
        return "credit_amount"
    if "贷方金额" in compact and "支" in compact:
        return "debit_amount"
    if compact in OWN_ACCOUNT_HEADERS or compact in ("本方账户", "企业账户"):
        return "account_number"
    for canonical, names in HEADER_SYNONYMS.items():
        if canonical in {"counterparty_account", "counterparty_bank", "account_number"}:
            continue
        if any(name in compact for name in names):
            return canonical
    if compact == "户名":
        return "counterparty_name"
    return None


def _find_header_row(rows: list[list[Any]], warnings: list[str], sheet_name: str) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_mapping: dict[int, str] = {}
    best_score = 0
    for index, row in enumerate(rows[:30]):
        mapping: dict[int, str] = {}
        for col_index, value in enumerate(row):
            canonical = _canonical_header(value)
            if canonical and canonical not in mapping.values():
                mapping[col_index] = canonical
        fields = set(mapping.values())
        amount_score = int("credit_amount" in fields) + int("debit_amount" in fields)
        score = amount_score * 2 + int("transaction_date" in fields) + int("balance" in fields) + int("summary" in fields)
        if score > best_score:
            best_score = score
            best_index = index
            best_mapping = mapping
        if score >= 4 and "transaction_date" in fields:
            return index, mapping
    if not best_mapping:
        warnings.append(f"sheet {sheet_name} 未能在前30行识别交易表头")
    return best_index, best_mapping


def _iter_label_value_pairs(rows: list[list[Any]], limit: int = 20) -> list[tuple[str, Any]]:
    pairs: list[tuple[str, Any]] = []
    for row in rows[:limit]:
        cells = [cell for cell in row]
        for index, cell in enumerate(cells):
            label = _compact(cell).rstrip(":：")
            if not label:
                continue
            value = cells[index + 1] if index + 1 < len(cells) else None
            if value not in (None, ""):
                pairs.append((label, value))
            text = normalize_text(cell)
            for sep in (":", "："):
                if sep in text:
                    left, right = text.split(sep, 1)
                    if left.strip() and right.strip():
                        pairs.append((_compact(left), right.strip()))
    return pairs


def _extract_account_from_column(rows: list[dict[str, Any]]) -> str | None:
    values = [normalize_account_number(row.get("account_number")) for row in rows]
    values = [value for value in values if value]
    if not values:
        return None
    return Counter(values).most_common(1)[0][0]


def _clean_enterprise_account(value: Any) -> str | None:
    text = normalize_text(value)
    text = re.sub(r"[（(]\s*人民币\s*[）)]", "", text)
    return normalize_account_number(text)


def _summary_label(label: str) -> str:
    text = _compact(label)
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\((?:¥|￥|元|人民币)?\)", "", text)
    text = text.replace("¥", "").replace("￥", "").replace("元", "")
    return text


def parse_sheet_account_info(sheet_name: str, rows: list[list[Any]], customer_name: str | None = None) -> dict[str, Any]:
    bank_name = infer_bank_from_sheet_name(sheet_name) or sheet_name
    account_info: dict[str, Any] = {
        "bank_name": bank_name,
        "account_name": customer_name,
        "account_number": None,
        "branch_name": None,
        "currency": "人民币",
        "period_start": None,
        "period_end": None,
        "summary_inflow": None,
        "summary_outflow": None,
        "summary_inflow_count": None,
        "summary_outflow_count": None,
    }

    for label, value in _iter_label_value_pairs(rows):
        normalized_label = _summary_label(label)
        if any(key in normalized_label for key in ("对方账号", "对方账户", "收付款方账号", "交易对手账号")):
            continue
        if any(key in normalized_label for key in ("对方开户行", "对方机构", "对方银行", "收付款方开户行")):
            continue
        if normalized_label in {"账户名称", "户名", "企业名称", "客户名称", "单位名称"}:
            account_info["account_name"] = normalize_text(value) or account_info["account_name"]
        elif normalized_label in {"账号", "企业账号", "本方账号", "账户账号", "银行账号", "账户号码"}:
            account_info["account_number"] = _clean_enterprise_account(value)
        elif normalized_label in {"开户机构", "开户行", "开户网点"}:
            account_info["branch_name"] = normalize_text(value) or None
        elif normalized_label in {"币种", "货币"}:
            account_info["currency"] = normalize_currency(value)
        elif normalized_label in {"起始日期", "开始日期", "流水开始日期"}:
            account_info["period_start"] = normalize_date(value)
        elif normalized_label in {"截止日期", "结束日期", "流水结束日期"}:
            account_info["period_end"] = normalize_date(value)
        elif normalized_label in {"借方累计发生额", "总支出", "贷方交易金额"}:
            account_info["summary_outflow"] = normalize_amount(value)
        elif normalized_label in {"贷方累计发生额", "总收入", "借方交易金额"}:
            account_info["summary_inflow"] = normalize_amount(value)
        elif normalized_label in {"借方累计笔数", "总支出笔数", "贷方交易笔数"}:
            amount = normalize_amount(value)
            account_info["summary_outflow_count"] = int(amount) if amount is not None else None
        elif normalized_label in {"贷方累计笔数", "总收入笔数", "借方交易笔数"}:
            amount = normalize_amount(value)
            account_info["summary_inflow_count"] = int(amount) if amount is not None else None

    account_id = f"{account_info['bank_name']}:{account_info.get('account_number') or sheet_name}"
    account_info["account_id"] = account_id
    return account_info


def read_excel_workbook(file_path: str | None = None, rows: list[dict[str, Any]] | None = None, filename: str | None = None) -> dict[str, Any]:
    warnings: list[str] = []
    sheets: list[dict[str, Any]] = []
    source_file = filename or (Path(file_path).name if file_path else "")
    if file_path:
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            warnings.append(f"{suffix or 'unknown'} 暂未直接读取 workbook，已尝试使用上传解析 rows fallback")
        else:
            try:
                workbook = load_workbook(path, read_only=True, data_only=True)
                for ws in workbook.worksheets:
                    raw_rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
                    header_index, header_map = _find_header_row(raw_rows, warnings, ws.title)
                    data_rows: list[dict[str, Any]] = []
                    detected_columns: list[str] = []
                    column_mapping: dict[str, str] = {}
                    if header_index is not None and header_map:
                        header_row = raw_rows[header_index]
                        detected_columns = [normalize_text(header_row[idx] or f"col_{idx + 1}") for idx in range(len(header_row))]
                        column_mapping = {
                            normalize_text(header_row[idx] or f"col_{idx + 1}"): field
                            for idx, field in header_map.items()
                        }
                        for row_number, raw_row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
                            raw = {str(header_row[idx] or f"col_{idx + 1}"): raw_row[idx] if idx < len(raw_row) else None for idx in range(len(header_row))}
                            normalized = {field: raw_row[idx] if idx < len(raw_row) else None for idx, field in header_map.items()}
                            normalized["raw"] = raw
                            normalized["row_number"] = row_number
                            data_rows.append(normalized)
                    meta = parse_sheet_account_info(ws.title, raw_rows)
                    debug = {
                        "header_index": header_index,
                        "detected_columns": detected_columns,
                        "column_mapping": column_mapping,
                        "header_summary": {
                            "total_inflow": meta.get("summary_inflow"),
                            "total_outflow": meta.get("summary_outflow"),
                            "inflow_count": meta.get("summary_inflow_count"),
                            "outflow_count": meta.get("summary_outflow_count"),
                        },
                    }
                    if "泰隆" in normalize_text(ws.title):
                        logger.info(
                            "[EnterpriseFlow][Tailong][COLUMNS] sheet=%s columns=%s mapping=%s",
                            ws.title,
                            detected_columns,
                            column_mapping,
                        )
                    if not meta.get("account_number"):
                        meta["account_number"] = _extract_account_from_column(data_rows)
                        meta["account_id"] = f"{meta['bank_name']}:{meta.get('account_number') or ws.title}"
                    sheets.append({"sheet_name": ws.title, "meta": meta, "rows": data_rows, "debug": debug})
                workbook.close()
                return {"source_file": source_file, "sheets": sheets, "warnings": warnings}
            except Exception as exc:
                warnings.append(f"Excel workbook 读取失败：{exc}")

    if rows:
        data_rows = []
        for index, row in enumerate(rows, start=2):
            normalized: dict[str, Any] = {}
            for key, value in row.items():
                canonical = _canonical_header(key)
                if canonical:
                    normalized[canonical] = value
            normalized["raw"] = dict(row)
            normalized["row_number"] = index
            data_rows.append(normalized)
        sheet_name = "上传解析结果"
        meta = parse_sheet_account_info(sheet_name, [list(item.keys()) for item in rows[:20]])
        meta["account_number"] = meta.get("account_number") or _extract_account_from_column(data_rows)
        meta["account_id"] = f"{meta['bank_name']}:{meta.get('account_number') or sheet_name}"
        sheets.append({"sheet_name": sheet_name, "meta": meta, "rows": data_rows})
    return {"source_file": source_file, "sheets": sheets, "warnings": warnings}


# ---------------------------------------------------------------------------
# Robust bank statement reader overrides.
# The original helpers above are kept for compatibility, but these definitions
# intentionally appear last so the Agent uses Chinese-header aware parsing.

GENERIC_SHEET_NAMES = {"testReport", "Sheet", "Sheet1", "明细", "账户明细查询", ""}

BANK_ALIASES_CN: tuple[tuple[str, str], ...] = (
    ("北京银行", "北京银行"),
    ("中国建设银行", "建设银行"),
    ("上海银行", "上海银行"),
    ("建行", "建设银行"),
    ("建设银行", "建设银行"),
    ("民生", "民生银行"),
    ("平安", "平安银行"),
    ("泰隆", "泰隆银行"),
    ("浙江网商", "浙江网商银行"),
    ("网商", "浙江网商银行"),
)

HEADER_SYNONYMS_CN: dict[str, tuple[str, ...]] = {
    "transaction_date": ("交易日期", "交易时间", "记账日期", "入账日期", "发生日期", "日期", "提交时间"),
    "post_date": ("记账日期", "入账日期", "过账日期"),
    "credit_amount": ("贷方发生额", "贷方金额", "贷方发生额（收入）", "贷方发生额(收入)", "收入", "收入金额", "入账金额", "转入", "发生额收入"),
    "debit_amount": ("借方发生额", "借方金额", "借方发生额（支取）", "借方发生额(支取)", "支出", "支出金额", "出账金额", "转出", "发生额支出"),
    "balance": ("余额", "账户余额", "交易后余额", "可用余额"),
    "summary": ("摘要", "交易摘要", "交易名称"),
    "purpose": ("交易用途", "用途"),
    "remark": ("备注", "附言", "客户附言"),
    "counterparty_name": ("对方户名", "对方名称", "对手名称", "交易对手", "收付款方名称"),
    "counterparty_account": ("对方账号", "对方账户", "对手账号", "收付款方账号", "交易对手账号"),
    "counterparty_bank": ("对方开户行", "对方机构", "对方银行", "收付款方开户行", "对方账户开户行"),
    "payee_name": ("收款人", "收款方", "付款人", "付款方"),
    "payee_account": ("收款账号", "收款账户", "付款账号", "付款账户"),
    "account_number": ("账号", "本方账号", "企业账号", "选择账号", "账户账号", "账户号码"),
    "account_name": ("账户名称", "户名", "企业名称", "客户名称", "单位名称"),
    "currency": ("币种", "货币"),
    "transaction_direction": ("交易方向",),
    "serial_no": ("交易流水号", "账户明细编号", "企业流水号", "凭证号"),
}

SHANGHAI_DETAIL_HEADERS = (
    "交易流水号",
    "交易时间",
    "记账日期",
    "交易方向",
    "借方发生额",
    "贷方发生额",
    "余额",
    "对手账号",
    "对手名称",
    "摘要",
    "交易用途",
    "备注",
)

TOP_KV_LABELS = (
    "借方总笔数",
    "贷方总笔数",
    "借方总金额",
    "贷方总金额",
    "选择账号",
    "选择帐号",
    "企业账号",
    "企业帐号",
    "记账日期",
    "开户行",
    "总笔数",
    "户名",
    "币种",
)


def _cn_text(value: Any) -> str:
    return normalize_text(value).replace("\u3000", " ").strip()


def _cn_compact(value: Any) -> str:
    return re.sub(r"\s+", "", _cn_text(value))


def _strip_label(value: Any) -> str:
    text = _cn_compact(value)
    text = text.rstrip(":：")
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\((?:¥|￥|元|人民币)\)", "", text)
    return text


def _split_compound_cell(value: Any) -> list[str]:
    text = _cn_text(value)
    if not text:
        return []
    # Some exports put the whole header row into one tab/newline/multi-space
    # cell. Split those first; if the known Shanghai headers are merely
    # concatenated, recover them in their expected order.
    parts = [part.strip() for part in re.split(r"[\t\r\n]+|\s{2,}", text) if part.strip()]
    if len(parts) > 1:
        return parts
    compact = _cn_compact(text)
    if all(token in compact for token in ("交易流水号", "交易时间", "借方发生额", "贷方发生额", "余额")):
        found = [token for token in SHANGHAI_DETAIL_HEADERS if token in compact]
        if len(found) >= 6:
            return found
    return [text]


def _expand_row_cells(row: list[Any]) -> list[Any]:
    non_empty = [_cn_text(cell) for cell in row if _cn_text(cell)]
    if len(non_empty) <= 2:
        expanded: list[Any] = []
        changed = False
        for cell in row:
            pieces = _split_compound_cell(cell)
            if len(pieces) > 1:
                changed = True
                expanded.extend(pieces)
            else:
                expanded.append(cell)
        if changed:
            return expanded
    return row


def _extract_inline_top_kv(text: str) -> list[tuple[str, str]]:
    normalized = _cn_text(text)
    if not normalized:
        return []
    pattern = r"(" + "|".join(re.escape(label) for label in TOP_KV_LABELS) + r")\s*[:：]?"
    matches = list(re.finditer(pattern, normalized))
    pairs: list[tuple[str, str]] = []
    for index, match in enumerate(matches):
        label = _strip_label(match.group(1))
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        value = normalized[start:end].strip(" ：:\t\r\n")
        if value:
            pairs.append((label, value))
    return pairs


def infer_bank_from_text(*values: Any) -> str | None:
    text = " ".join(_cn_text(value) for value in values if value)
    for keyword, bank_name in BANK_ALIASES_CN:
        if keyword in text:
            return bank_name
    if "银行" in text and text not in GENERIC_SHEET_NAMES:
        match = re.search(r"([\u4e00-\u9fa5]{2,20}银行)", text)
        if match:
            return match.group(1)
    return None


def infer_bank_from_sheet_name(sheet_name: Any) -> str | None:
    text = _cn_text(sheet_name)
    if text in GENERIC_SHEET_NAMES:
        return None
    return infer_bank_from_text(text)


def _canonical_header(value: Any) -> str | None:
    label = _strip_label(value)
    if not label:
        return None
    if any(token in label for token in ("收款账号", "收款账户", "付款账号", "付款账户")):
        return "payee_account"
    if any(token in label for token in ("收款人", "收款方", "付款人", "付款方")):
        return "payee_name"
    if label == "对手方":
        return "counterparty_name"
    if "对手方账号" in label:
        return "counterparty_account"
    if "对手方银行" in label:
        return "counterparty_bank"
    # Shanghai Bank "账户明细查询" exports use very stable Chinese headers.
    # Check these first so "对手账号" never falls through to own account number.
    if "对手账号" in label or "对方账号" in label or "交易对手账号" in label:
        return "counterparty_account"
    if "对手名称" in label or "对方户名" in label or "对方名称" in label:
        return "counterparty_name"
    if "借方发生额" in label or "借方金额" in label:
        return "debit_amount"
    if "贷方发生额" in label or "贷方金额" in label:
        return "credit_amount"
    if "交易流水号" in label:
        return "serial_no"
    if label == "交易时间":
        return "transaction_date"
    if label == "记账日期":
        return "post_date"
    if label == "交易方向":
        return "transaction_direction"
    if any(name in label for name in HEADER_SYNONYMS_CN["counterparty_bank"]):
        return "counterparty_bank"
    if any(name in label for name in HEADER_SYNONYMS_CN["counterparty_account"]):
        return "counterparty_account"
    for canonical, names in HEADER_SYNONYMS_CN.items():
        if canonical in {"counterparty_bank", "counterparty_account", "payee_name", "payee_account"}:
            continue
        if any(name in label for name in names):
            if canonical == "account_number" and any(token in label for token in ("对方", "对手", "收付款方", "交易对手")):
                return "counterparty_account"
            return canonical
    return None


def _find_header_row(rows: list[list[Any]], warnings: list[str], sheet_name: str) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_mapping: dict[int, str] = {}
    best_score = 0
    candidate_rows: list[dict[str, Any]] = []
    for index, row in enumerate(rows[:50]):
        row = _expand_row_cells(list(row))
        mapping: dict[int, str] = {}
        seen: set[str] = set()
        for col_index, value in enumerate(row):
            canonical = _canonical_header(value)
            if canonical and canonical not in seen:
                mapping[col_index] = canonical
                seen.add(canonical)
        fields = set(mapping.values())
        score = 0
        row_text = _cn_compact(" ".join(_cn_text(cell) for cell in row if _cn_text(cell)))
        is_shanghai_detail_header = all(
            token in row_text
            for token in ("交易流水号", "交易时间", "借方发生额", "贷方发生额", "余额")
        ) and ("对手名称" in row_text or "对方名称" in row_text)
        if is_shanghai_detail_header:
            score = 12
        elif {"serial_no", "transaction_date", "debit_amount", "credit_amount"}.issubset(fields):
            score = 10
        elif {"account_number", "account_name", "transaction_date", "debit_amount", "credit_amount"}.issubset(fields):
            score = 10
        elif {"transaction_date", "credit_amount", "debit_amount", "balance"}.issubset(fields):
            score = 9
        elif {"debit_amount", "credit_amount", "balance"}.issubset(fields):
            score = 8
        else:
            score = (
                int("transaction_date" in fields)
                + int("debit_amount" in fields) * 2
                + int("credit_amount" in fields) * 2
                + int("balance" in fields)
                + int("account_number" in fields)
                + int("account_name" in fields)
            )
        if score > best_score:
            best_score = score
            best_index = index
            best_mapping = mapping
        if score > 0:
            candidate_rows.append({
                "row": index + 1,
                "score": score,
                "fields": sorted(fields),
                "text": row_text[:300],
            })
        if score >= 8:
            logger.info("[EnterpriseFlow][HeaderScan] sheet=%s candidate_rows=%s", sheet_name, candidate_rows[:8])
            logger.info(
                "[EnterpriseFlow][HeaderDetect] sheet=%s header_row=%s columns=%s",
                sheet_name,
                index + 1,
                [_cn_text(cell) for cell in row],
            )
            return index, mapping
    if best_mapping:
        logger.info("[EnterpriseFlow][HeaderScan] sheet=%s candidate_rows=%s", sheet_name, candidate_rows[:8])
        logger.info(
            "[EnterpriseFlow][HeaderDetect] sheet=%s header_row=%s columns=%s",
            sheet_name,
            (best_index or 0) + 1,
            [_cn_text(cell) for cell in _expand_row_cells(list(rows[best_index or 0]))],
        )
    else:
        sample_rows = [
            " | ".join(_cn_text(cell) for cell in _expand_row_cells(list(row)) if _cn_text(cell))[:500]
            for row in rows[:10]
        ]
        logger.warning(
            "[EnterpriseFlow][HeaderDetectFailed] sheet=%s sample_rows=%s",
            sheet_name,
            sample_rows,
        )
        warnings.append(f"sheet {sheet_name} 未识别到交易表头")
    return best_index, best_mapping


def _iter_label_value_pairs(rows: list[list[Any]], limit: int = 30, sheet_name: str | None = None) -> list[tuple[str, Any]]:
    pairs: list[tuple[str, Any]] = []
    for row in rows[:limit]:
        cells = list(_expand_row_cells(list(row))[:15])
        joined_text = " ".join(_cn_text(cell) for cell in cells if _cn_text(cell))
        pairs.extend(_extract_inline_top_kv(joined_text))
        header_like_count = sum(1 for cell in cells if _canonical_header(cell))
        if header_like_count >= 4:
            continue
        for index, cell in enumerate(cells):
            raw_text = _cn_text(cell)
            label = _strip_label(raw_text)
            if not label:
                continue
            for sep in (":", "："):
                if sep in raw_text:
                    left, right = raw_text.split(sep, 1)
                    if left.strip() and right.strip():
                        pairs.append((_strip_label(left), right.strip()))
            next_value = cells[index + 1] if index + 1 < len(cells) else None
            if next_value not in (None, ""):
                pairs.append((label, next_value))
            else:
                # Some bank exports place key/value pairs as "选择账号 | 空列 | 值"
                # or "开户行 | 空列 | 值 | 币种 | 值". Look ahead a few cells,
                # stopping when another label-like cell begins.
                for offset in range(2, 6):
                    candidate_index = index + offset
                    if candidate_index >= len(cells):
                        break
                    candidate = cells[candidate_index]
                    if candidate in (None, ""):
                        continue
                    candidate_text = _cn_text(candidate)
                    if _canonical_header(candidate_text) or _strip_label(candidate_text) in {
                        "选择账号", "户名", "开户行", "币种", "总笔数", "借方总笔数",
                        "贷方总笔数", "借方总金额", "贷方总金额", "记账日期",
                    }:
                        break
                    pairs.append((label, candidate))
                    break
    found_keys = []
    for label, value in pairs:
        key = _strip_label(label)
        if key and value not in (None, "") and key not in found_keys:
            found_keys.append(key)
    logger.info("[EnterpriseFlow][TopKVScan] sheet=%s found_keys=%s", sheet_name or "", found_keys)
    return pairs


def _parse_date_range(value: Any) -> tuple[str | None, str | None]:
    text = _cn_text(value)
    parts = re.split(r"\s*(?:---|--|至|~|～)\s*", text)
    if len(parts) >= 2:
        return normalize_date(parts[0]), normalize_date(parts[1])
    return normalize_date(text), None


def parse_sheet_account_info(sheet_name: str, rows: list[list[Any]], customer_name: str | None = None, source_file: str | None = None) -> dict[str, Any]:
    raw_bank_context = " ".join(_cn_text(item) for item in (source_file, sheet_name) if item)
    bank_name = infer_bank_from_text(source_file) or infer_bank_from_sheet_name(sheet_name)
    logger.info("[EnterpriseFlow][BankDetect] file=%s raw=%s normalized_bank=%s", source_file or "", raw_bank_context, bank_name or "")
    account_info: dict[str, Any] = {
        "bank_name": bank_name or "未知银行",
        "account_name": customer_name,
        "account_number": None,
        "branch_name": None,
        "currency": "人民币",
        "period_start": None,
        "period_end": None,
        "summary_inflow": None,
        "summary_outflow": None,
        "summary_inflow_count": None,
        "summary_outflow_count": None,
        "summary_transaction_count": None,
    }

    for label, value in _iter_label_value_pairs(rows, sheet_name=sheet_name):
        label = _strip_label(label)
        if any(token in label for token in ("对方账号", "对手账号", "收付款方账号", "交易对手账号", "对方开户行", "对方银行")):
            continue
        if label in {"账户名称", "户名", "企业名称", "客户名称", "单位名称"}:
            account_info["account_name"] = _cn_text(value) or account_info["account_name"]
        elif label in {"账号", "企业账号", "企业帐号", "本方账号", "选择账号", "选择帐号", "账户账号", "账户号码"}:
            account_info["account_number"] = _clean_enterprise_account(value)
        elif label in {"开户行", "开户机构", "开户网点"}:
            branch = _cn_text(value)
            account_info["branch_name"] = branch or None
            account_info["bank_name"] = infer_bank_from_text(branch) or account_info["bank_name"]
        elif label in {"币种", "货币"}:
            account_info["currency"] = normalize_currency(value)
        elif label in {"记账日期", "起始日期", "开始日期", "流水开始日期"}:
            start, end = _parse_date_range(value)
            account_info["period_start"] = start or account_info["period_start"]
            account_info["period_end"] = end or account_info["period_end"]
        elif label in {"截止日期", "结束日期", "流水结束日期"}:
            account_info["period_end"] = normalize_date(value)
        elif "总笔数" == label:
            amount = normalize_amount(value)
            account_info["summary_transaction_count"] = int(amount) if amount is not None else None
        elif "借方总笔数" in label or label in {"借方累计笔数", "总支出笔数", "贷方交易笔数"}:
            amount = normalize_amount(value)
            account_info["summary_outflow_count"] = int(amount) if amount is not None else None
        elif "贷方总笔数" in label or label in {"贷方累计笔数", "总收入笔数", "借方交易笔数"}:
            amount = normalize_amount(value)
            account_info["summary_inflow_count"] = int(amount) if amount is not None else None
        elif "借方总金额" in label or label in {"借方累计发生额", "总支出", "贷方交易金额"}:
            account_info["summary_outflow"] = normalize_amount(value)
        elif "贷方总金额" in label or label in {"贷方累计发生额", "总收入", "借方交易金额"}:
            account_info["summary_inflow"] = normalize_amount(value)

    if account_info["bank_name"] == "未知银行":
        account_info["bank_name"] = infer_bank_from_text(sheet_name) or "未知银行"
    account_id = f"{account_info['bank_name']}:{account_info.get('account_number') or source_file or sheet_name}"
    account_info["account_id"] = account_id
    return account_info


def _sheet_raw_stats(rows: list[list[Any]]) -> tuple[int, int]:
    non_empty_cells = 0
    non_empty_rows = 0
    for row in rows:
        row_has_value = False
        for cell in row:
            if _cn_text(cell):
                non_empty_cells += 1
                row_has_value = True
        if row_has_value:
            non_empty_rows += 1
    return non_empty_cells, non_empty_rows


def _sheet_preview(rows: list[list[Any]], limit: int = 15) -> list[str]:
    preview: list[str] = []
    for row in rows[:limit]:
        text = " | ".join(_cn_text(cell) for cell in _expand_row_cells(list(row)) if _cn_text(cell))
        preview.append(text)
    return preview


def _log_sheet_raw_diagnostics(
    sheet_name: str,
    rows: list[list[Any]],
    *,
    max_row: int | None = None,
    max_column: int | None = None,
    merged_ranges: int = 0,
    strategy: str = "",
) -> None:
    non_empty_cells, non_empty_rows = _sheet_raw_stats(rows)
    logger.info(
        "[EnterpriseFlow][SheetRawStats] sheet=%s strategy=%s max_row=%s max_column=%s non_empty_cells=%s non_empty_rows=%s merged_ranges=%s",
        sheet_name,
        strategy,
        max_row if max_row is not None else len(rows),
        max_column if max_column is not None else max((len(row) for row in rows), default=0),
        non_empty_cells,
        non_empty_rows,
        merged_ranges,
    )
    logger.info("[EnterpriseFlow][SheetRawPreview] sheet=%s rows=%s", sheet_name, _sheet_preview(rows))


def _worksheet_rows_with_merged_values(ws: Any) -> list[list[Any]]:
    max_row = int(getattr(ws, "max_row", 0) or 0)
    max_column = int(getattr(ws, "max_column", 0) or 0)
    if max_row <= 0 or max_column <= 0:
        return []
    merged_lookup: dict[tuple[int, int], Any] = {}
    for merged_range in getattr(ws, "merged_cells", ()).ranges:
        min_col, min_row, max_col, max_merged_row = range_boundaries(str(merged_range))
        top_left_value = ws.cell(min_row, min_col).value
        for row_index in range(min_row, max_merged_row + 1):
            for col_index in range(min_col, max_col + 1):
                merged_lookup[(row_index, col_index)] = top_left_value

    rows: list[list[Any]] = []
    for row_index in range(1, max_row + 1):
        row: list[Any] = []
        for col_index in range(1, max_column + 1):
            value = ws.cell(row_index, col_index).value
            if value in (None, "") and (row_index, col_index) in merged_lookup:
                value = merged_lookup[(row_index, col_index)]
            row.append(value)
        rows.append(row)
    return rows


def _read_xlsx_rows_openpyxl(path: Path, *, data_only: bool, strategy: str) -> list[tuple[str, list[list[Any]]]]:
    workbook = load_workbook(path, read_only=False, data_only=data_only)
    result: list[tuple[str, list[list[Any]]]] = []
    try:
        for ws in workbook.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            rows = _worksheet_rows_with_merged_values(ws)
            merged_ranges = len(list(getattr(ws, "merged_cells", ()).ranges))
            _log_sheet_raw_diagnostics(
                ws.title,
                rows,
                max_row=getattr(ws, "max_row", None),
                max_column=getattr(ws, "max_column", None),
                merged_ranges=merged_ranges,
                strategy=strategy,
            )
            result.append((ws.title, rows))
        return result
    finally:
        workbook.close()


def _read_xlsx_rows_pandas(path: Path) -> list[tuple[str, list[list[Any]]]]:
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise RuntimeError("pandas fallback unavailable") from exc
    sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=object, engine="openpyxl")
    result: list[tuple[str, list[list[Any]]]] = []
    for sheet_name, frame in sheets.items():
        frame = frame.where(frame.notna(), None)
        rows = frame.values.tolist()
        _log_sheet_raw_diagnostics(sheet_name, rows, strategy="pandas-openpyxl")
        result.append((str(sheet_name), rows))
    return result


def _total_non_empty_cells(raw_sheets: list[tuple[str, list[list[Any]]]]) -> int:
    return sum(_sheet_raw_stats(rows)[0] for _, rows in raw_sheets)


def _read_xlsx_rows(path: Path) -> list[tuple[str, list[list[Any]]]]:
    attempts: list[tuple[str, Any]] = [
        ("openpyxl-data-only", lambda: _read_xlsx_rows_openpyxl(path, data_only=True, strategy="openpyxl-data-only")),
        ("openpyxl-formula", lambda: _read_xlsx_rows_openpyxl(path, data_only=False, strategy="openpyxl-formula")),
        ("pandas-openpyxl", lambda: _read_xlsx_rows_pandas(path)),
        ("html-fallback", lambda: _read_html_xls_rows(path)),
    ]
    errors: list[str] = []
    best: list[tuple[str, list[list[Any]]]] = []
    best_cells = 0
    for name, reader in attempts:
        try:
            candidate = reader()
            cells = _total_non_empty_cells(candidate)
            if cells > best_cells:
                best = candidate
                best_cells = cells
            if cells >= 5:
                return candidate
            logger.warning("[EnterpriseFlow][WorkbookReadEmpty] file=%s strategy=%s non_empty_cells=%s", path.name, name, cells)
        except Exception as exc:
            errors.append(f"{name}: {exc}")
            logger.warning("[EnterpriseFlow][WorkbookReadFailed] file=%s strategy=%s reason=%s", path.name, name, exc)
    if best_cells == 0 and errors:
        logger.warning("[EnterpriseFlow][WorkbookReadFailed] file=%s strategies=%s", path.name, errors)
    return best


def _read_xls_rows(path: Path) -> list[tuple[str, list[list[Any]]]]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise RuntimeError("读取 .xls 需要安装 xlrd>=2.0.1") from exc
    book = xlrd.open_workbook(str(path))
    result: list[tuple[str, list[list[Any]]]] = []
    for sheet in book.sheets():
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        _log_sheet_raw_diagnostics(
            sheet.name,
            rows,
            max_row=sheet.nrows,
            max_column=sheet.ncols,
            strategy="xlrd",
        )
        result.append((sheet.name, rows))
    return result


def _read_html_xls_rows(path: Path) -> list[tuple[str, list[list[Any]]]]:
    raw = path.read_bytes()
    text = ""
    for encoding in ("utf-8", "gb18030", "gbk"):
        try:
            text = raw.decode(encoding)
            break
        except Exception:
            continue
    if not text:
        text = raw.decode("utf-8", errors="ignore")
    tables = re.findall(r"<table[\s\S]*?</table>", text, flags=re.IGNORECASE)
    result: list[tuple[str, list[list[Any]]]] = []
    for idx, table in enumerate(tables, start=1):
        rows: list[list[Any]] = []
        for tr in re.findall(r"<tr[\s\S]*?</tr>", table, flags=re.IGNORECASE):
            cells = re.findall(r"<t[dh][^>]*>([\s\S]*?)</t[dh]>", tr, flags=re.IGNORECASE)
            clean_cells = [re.sub(r"<[^>]+>", "", cell).replace("&nbsp;", " ").strip() for cell in cells]
            if any(clean_cells):
                rows.append(clean_cells)
        if rows:
            sheet_name = f"sheet{idx}"
            _log_sheet_raw_diagnostics(sheet_name, rows, strategy="html-table")
            result.append((sheet_name, rows))
    if result:
        return result

    worksheets = re.findall(
        r"<(?:\w+:)?Worksheet\b[^>]*(?:ss:)?Name=[\"']([^\"']+)[\"'][^>]*>([\s\S]*?)</(?:\w+:)?Worksheet>",
        text,
        flags=re.IGNORECASE,
    )
    for idx, (name, body) in enumerate(worksheets, start=1):
        rows = []
        for row_xml in re.findall(r"<(?:\w+:)?Row\b[^>]*>([\s\S]*?)</(?:\w+:)?Row>", body, flags=re.IGNORECASE):
            cells = []
            for cell_xml in re.findall(r"<(?:\w+:)?Cell\b[^>]*>([\s\S]*?)</(?:\w+:)?Cell>", row_xml, flags=re.IGNORECASE):
                data_match = re.search(r"<(?:\w+:)?Data\b[^>]*>([\s\S]*?)</(?:\w+:)?Data>", cell_xml, flags=re.IGNORECASE)
                value = data_match.group(1) if data_match else cell_xml
                value = re.sub(r"<[^>]+>", "", value).replace("&nbsp;", " ").strip()
                cells.append(value)
            if any(cells):
                rows.append(cells)
        if rows:
            sheet_name = re.sub(r"<[^>]+>", "", name).strip() or f"sheet{idx}"
            _log_sheet_raw_diagnostics(sheet_name, rows, strategy="spreadsheetml")
            result.append((sheet_name, rows))
    if result:
        return result

    if "," in text or "\t" in text:
        delimiter = "\t" if "\t" in text else ","
        rows = [
            [cell.strip().strip('"') for cell in line.split(delimiter)]
            for line in text.splitlines()
            if line.strip()
        ]
        if rows:
            _log_sheet_raw_diagnostics("text-table", rows, strategy="text-delimited")
            result.append(("text-table", rows))
    return result


def _build_sheet(sheet_name: str, raw_rows: list[list[Any]], source_file: str, warnings: list[str]) -> dict[str, Any]:
    expanded_rows = [_expand_row_cells(list(row)) for row in raw_rows]
    header_index, header_map = _find_header_row(expanded_rows, warnings, sheet_name)
    data_rows: list[dict[str, Any]] = []
    detected_columns: list[str] = []
    column_mapping: dict[str, str] = {}
    if header_index is not None and header_map:
        header_row = expanded_rows[header_index]
        detected_columns = [_cn_text(header_row[idx] if idx < len(header_row) else f"col_{idx + 1}") for idx in range(len(header_row))]
        column_mapping = {
            _cn_text(header_row[idx] if idx < len(header_row) else f"col_{idx + 1}"): field
            for idx, field in header_map.items()
        }
        empty_streak = 0
        for row_number, raw_row in enumerate(expanded_rows[header_index + 1 :], start=header_index + 2):
            if not any(_cn_text(cell) for cell in raw_row):
                empty_streak += 1
                if empty_streak >= 8:
                    break
                continue
            empty_streak = 0
            raw = {str(header_row[idx] if idx < len(header_row) and header_row[idx] not in (None, "") else f"col_{idx + 1}"): raw_row[idx] if idx < len(raw_row) else None for idx in range(len(header_row))}
            normalized = {field: raw_row[idx] if idx < len(raw_row) else None for idx, field in header_map.items()}
            normalized["raw"] = raw
            normalized["row_number"] = row_number
            data_rows.append(normalized)
    meta = parse_sheet_account_info(sheet_name, expanded_rows, source_file=source_file)
    if not meta.get("account_number"):
        meta["account_number"] = _extract_account_from_column(data_rows)
        meta["account_id"] = f"{meta['bank_name']}:{meta.get('account_number') or source_file or sheet_name}"
    if not meta.get("account_name"):
        account_names = [_cn_text(row.get("account_name")) for row in data_rows if _cn_text(row.get("account_name"))]
        if account_names:
            meta["account_name"] = Counter(account_names).most_common(1)[0][0]
    if not meta.get("currency") or meta.get("currency") == "人民币":
        currencies = [_cn_text(row.get("currency")) for row in data_rows if _cn_text(row.get("currency"))]
        if currencies:
            meta["currency"] = normalize_currency(Counter(currencies).most_common(1)[0][0])
    logger.info(
        "[EnterpriseFlow][AccountInfo] sheet=%s bank=%s account_number=%s account_name=%s",
        sheet_name,
        meta.get("bank_name") or "",
        meta.get("account_number") or "",
        meta.get("account_name") or "",
    )
    debug = {
        "header_index": header_index,
        "detected_columns": detected_columns,
        "column_mapping": column_mapping,
        "header_summary": {
            "total_inflow": meta.get("summary_inflow"),
            "total_outflow": meta.get("summary_outflow"),
            "inflow_count": meta.get("summary_inflow_count"),
            "outflow_count": meta.get("summary_outflow_count"),
        },
    }
    return {"sheet_name": sheet_name, "meta": meta, "rows": data_rows, "debug": debug}


def read_excel_workbook(file_path: str | None = None, rows: list[dict[str, Any]] | None = None, filename: str | None = None) -> dict[str, Any]:
    warnings: list[str] = []
    sheets: list[dict[str, Any]] = []
    source_file = filename or (Path(file_path).name if file_path else "")
    if file_path:
        path = Path(file_path)
        suffix = path.suffix.lower()
        raw_sheets: list[tuple[str, list[list[Any]]]] = []
        try:
            if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                raw_sheets = _read_xlsx_rows(path)
            elif suffix == ".xls":
                try:
                    raw_sheets = _read_xls_rows(path)
                except Exception:
                    raw_sheets = _read_html_xls_rows(path)
                    if not raw_sheets:
                        raise
            else:
                warnings.append(f"{suffix or 'unknown'} 暂未支持直接读取 workbook")
        except Exception as exc:
            logger.warning("[EnterpriseFlow][WorkbookReadFailed] file=%s ext=%s reason=%s", source_file, suffix, exc)
            warnings.append(f"Excel workbook 读取失败：{exc}")
        if raw_sheets and _total_non_empty_cells(raw_sheets) == 0:
            warnings.append("Excel读取为空或未识别到交易表头")
        for sheet_name, raw_rows in raw_sheets:
            sheets.append(_build_sheet(sheet_name, raw_rows, source_file, warnings))
        return {"source_file": source_file, "sheets": sheets, "warnings": warnings}

    if rows:
        data_rows = []
        for index, row in enumerate(rows, start=2):
            normalized: dict[str, Any] = {}
            for key, value in row.items():
                canonical = _canonical_header(key)
                if canonical:
                    normalized[canonical] = value
            normalized["raw"] = dict(row)
            normalized["row_number"] = index
            data_rows.append(normalized)
        sheet_name = "上传解析结果"
        meta = parse_sheet_account_info(sheet_name, [list(item.keys()) for item in rows[:30]], source_file=source_file)
        meta["account_number"] = meta.get("account_number") or _extract_account_from_column(data_rows)
        meta["account_id"] = f"{meta['bank_name']}:{meta.get('account_number') or source_file or sheet_name}"
        sheets.append({"sheet_name": sheet_name, "meta": meta, "rows": data_rows, "debug": {}})
    return {"source_file": source_file, "sheets": sheets, "warnings": warnings}
