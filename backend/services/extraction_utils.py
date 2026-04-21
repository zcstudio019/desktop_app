from __future__ import annotations

import io
import math
import re
from collections import Counter
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

try:
    import pandas as pd
except ImportError:  # pragma: no cover - optional dependency fallback
    pd = None


SCIENTIFIC_NOTATION_RE = re.compile(r"^[+-]?\d+(?:\.\d+)?[eE][+-]?\d+$")
ACCOUNT_FIELD_KEYWORDS = (
    "账号",
    "账户",
    "卡号",
    "流水号",
    "交易号",
    "凭证号",
    "证件号",
    "身份证号",
    "统一社会信用代码",
    "信用代码",
    "编号",
    "开户许可证核准号",
    "核准号",
)
AMOUNT_FIELD_KEYWORDS = (
    "金额",
    "发生额",
    "余额",
    "收入",
    "支出",
    "贷方",
    "借方",
    "credit",
    "debit",
    "amount",
    "balance",
)
HEADER_HINT_KEYWORDS = (
    "日期",
    "交易日期",
    "记账日期",
    "摘要",
    "对方",
    "对手方",
    "收入",
    "支出",
    "余额",
    "借方",
    "贷方",
    "金额",
    "账号",
    "户名",
)


def normalize_text(value: Any) -> str:
    """Normalize cell/text values into clean strings."""
    if value is None:
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in {"nan", "none", "null", "nat"}:
            return ""
        return stripped
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null", "nat"}:
        return ""
    return text


def _scientific_to_plain(value: str) -> str:
    if not SCIENTIFIC_NOTATION_RE.match(value):
        return value
    try:
        decimal_value = Decimal(value)
    except InvalidOperation:
        return value
    plain = format(decimal_value, "f")
    if "." in plain:
        plain = plain.rstrip("0").rstrip(".")
    return plain


def normalize_amount(value: Any) -> str:
    """Normalize amount-like values into plain strings without scientific notation."""
    text = normalize_text(value)
    if not text:
        return ""

    text = text.replace(",", "").replace("，", "").replace(" ", "")
    text = re.sub(r"^[¥￥$]", "", text)
    text = _scientific_to_plain(text)
    if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        return ""
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def is_likely_account_field(field_name: str) -> bool:
    normalized = normalize_text(field_name)
    return any(keyword in normalized for keyword in ACCOUNT_FIELD_KEYWORDS)


def is_likely_amount_field(field_name: str) -> bool:
    normalized = normalize_text(field_name).lower()
    return any(keyword.lower() in normalized for keyword in AMOUNT_FIELD_KEYWORDS)


def _normalize_identifier(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = text.replace(" ", "")
    return _scientific_to_plain(text)


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: Counter[str] = Counter()
    result: list[str] = []
    for index, header in enumerate(headers):
        cleaned = normalize_text(header) or f"列{index + 1}"
        seen[cleaned] += 1
        if seen[cleaned] > 1:
            cleaned = f"{cleaned}_{seen[cleaned]}"
        result.append(cleaned)
    return result


def _score_header_row(cells: list[str]) -> int:
    score = 0
    for cell in cells:
        if not cell:
            continue
        if any(keyword in cell for keyword in HEADER_HINT_KEYWORDS):
            score += 3
        if len(cell) <= 12:
            score += 1
    return score


def _choose_header_index(rows: list[list[str]]) -> int:
    candidates = rows[: min(len(rows), 12)]
    best_index = 0
    best_score = -1
    for index, row in enumerate(candidates):
        score = _score_header_row(row)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def _read_with_pandas(file_source: str | bytes | io.BytesIO) -> list[dict[str, str]]:
    if pd is None:
        return []

    excel_source = io.BytesIO(file_source) if isinstance(file_source, bytes) else file_source
    sheet_map = pd.read_excel(excel_source, sheet_name=None, header=None, dtype=str)
    rows: list[dict[str, str]] = []
    for sheet_name, frame in sheet_map.items():
        raw_rows = [
            [normalize_text(value) for value in row]
            for row in frame.fillna("").values.tolist()
        ]
        raw_rows = [row for row in raw_rows if any(cell for cell in row)]
        if not raw_rows:
            continue
        header_index = _choose_header_index(raw_rows)
        headers = _dedupe_headers(raw_rows[header_index])
        for row in raw_rows[header_index + 1 :]:
            normalized_row: dict[str, str] = {"_sheet_name": sheet_name}
            has_content = False
            for idx, header in enumerate(headers):
                cell = row[idx] if idx < len(row) else ""
                value = _normalize_identifier(cell) if is_likely_account_field(header) else (
                    normalize_amount(cell) if is_likely_amount_field(header) else normalize_text(cell)
                )
                normalized_row[header] = value
                if value:
                    has_content = True
            if has_content:
                rows.append(normalized_row)
    return rows


def _read_with_openpyxl(file_source: str | bytes | io.BytesIO) -> list[dict[str, str]]:
    workbook_source = io.BytesIO(file_source) if isinstance(file_source, bytes) else file_source
    workbook = load_workbook(workbook_source, data_only=False)
    rows: list[dict[str, str]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        raw_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            normalized_row = [normalize_text(cell) for cell in row]
            if any(normalized_row):
                raw_rows.append(normalized_row)
        if not raw_rows:
            continue
        header_index = _choose_header_index(raw_rows)
        headers = _dedupe_headers(raw_rows[header_index])
        for row in raw_rows[header_index + 1 :]:
            normalized: dict[str, str] = {"_sheet_name": sheet_name}
            has_content = False
            for idx, header in enumerate(headers):
                cell = row[idx] if idx < len(row) else ""
                value = _normalize_identifier(cell) if is_likely_account_field(header) else (
                    normalize_amount(cell) if is_likely_amount_field(header) else normalize_text(cell)
                )
                normalized[header] = value
                if value:
                    has_content = True
            if has_content:
                rows.append(normalized)
    return rows


def read_excel_as_rows(file_path: str | bytes | io.BytesIO) -> list[dict]:
    """
    读取 xlsx/xls，返回逐行字典列表。
    所有单元格先按字符串读取并清洗，避免科学计数法。
    """
    try:
        rows = _read_with_pandas(file_path)
        if rows:
            return rows
    except Exception:
        pass

    return _read_with_openpyxl(file_path)


def rows_to_text(rows: list[dict[str, Any]], *, max_rows: int = 120) -> str:
    """Convert parsed Excel rows into compact text summaries without flattening the whole sheet."""
    lines: list[str] = []
    for row in rows[:max_rows]:
        parts = []
        for key, value in row.items():
            if key.startswith("_"):
                continue
            cleaned = normalize_text(value)
            if cleaned:
                parts.append(f"{key}: {cleaned}")
        if parts:
            lines.append(" | ".join(parts))
    return "\n".join(lines)

