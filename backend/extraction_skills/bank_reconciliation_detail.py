from __future__ import annotations

import csv
import hashlib
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .base import BaseExtractionSkill, ExtractionInput, ExtractionResult

logger = logging.getLogger(__name__)

DOC_TYPE = "bank_reconciliation_detail"
DOC_TYPE_NAME = "银行对账明细"
AGENT_TYPE = "bank_reconciliation_detail_agent"
SKILL_NAME = "bank_reconciliation_detail_skill"
SCHEMA_VERSION = "bank_reconciliation_detail.agent.v1"

PLACEHOLDER_VALUES = {"17"}
UNKNOWN = "未识别"

OPERATING_IN_KEYWORDS = ("工程款", "项目款", "货款", "材料款", "服务费", "劳务款", "咨询费", "结算款", "回款", "合同款", "进度款", "施工款", "设计费", "监理费")
OPERATING_OUT_KEYWORDS = ("材料款", "工程款", "项目款", "货款", "劳务费", "服务费", "采购款", "设备款", "安装费", "施工费", "分包款", "电缆款", "风管材料款", "灯具款", "防火包裹", "空调设备", "建材款")
NON_OPERATING_KEYWORDS = ("借款", "还款", "退款", "保证金", "押金", "备用金", "代垫款", "临时款", "内部款")
LOAN_KEYWORDS = ("融资租赁", "普惠融资", "贷款", "放款", "还款", "贷款利息", "利息支付", "对公贷款记账", "租赁", "担保", "小企业贷款", "小企业其他短期贷款利息收入")
FEE_KEYWORDS = ("手续费", "短信业务服务费", "工本费", "年费", "证书收入", "证书工本费", "账户服务费", "普通卡年费", "跨行汇款手续费", "到账伴侣")
SALARY_KEYWORDS = ("工资", "代发", "薪酬", "社保", "公积金")
TAX_KEYWORDS = ("税", "税款", "税务", "国库", "社保", "公积金")
DEPOSIT_KEYWORDS = ("押金", "保证金", "退回")
INTEREST_KEYWORDS = ("利息", "结息")
UNKNOWN_COUNTERPARTY_NAMES = {"", "未识别", "未知", "空", "none", "null", "None", "NULL"}
ORGANIZATION_KEYWORDS = ("公司", "银行", "账户", "中心", "集团", "有限", "合作社", "商行", "个体工商户", "厂", "店", "部", "局", "院", "所")
NOISE_COUNTERPARTY_KEYWORDS = (
    "代发专用账户",
    "贷款利息收入",
    "小企业其他短期贷款利息收入",
    "对公工行证书收入",
    "对公客户证书工本费收入",
    "网银注册账户服务费",
    "财智账户卡普通卡年费",
    "跨行汇款手续费",
    "手续费",
    "利息",
    "到账伴侣",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).replace("\u3000", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return "" if text in PLACEHOLDER_VALUES else text


def _compact(value: Any) -> str:
    return re.sub(r"\s+", "", _text(value))


def _normalize_party_name(value: Any) -> str:
    text = _text(value)
    if not text:
        return ""
    text = (
        text.replace("\u3000", "")
        .replace("\ufeff", "")
        .replace("\u200b", "")
        .replace("\u200c", "")
        .replace("\u200d", "")
        .replace("（", "(")
        .replace("）", ")")
    )
    return re.sub(r"[\s\r\n\t]+", "", text)


SELF_PARTY_KEYS = {
    "customer_name",
    "customerName",
    "company_name",
    "companyName",
    "name",
    "account_name",
    "accountName",
    "account_holder",
    "accountHolder",
    "客户名称",
    "企业名称",
    "公司名称",
    "户名",
}
CUSTOMER_CONTAINER_KEYS = {"customer", "customer_info", "customerInfo", "customer_profile", "customerProfile", "客户档案", "客户信息"}
LEGAL_REP_KEYS = {"legal_representative", "legalRepresentative", "legal_representative_name", "legalRepresentativeName", "legal_person", "legalPerson", "legal_person_name", "legalPersonName", "法人", "法人姓名", "法定代表人", "法定代表人姓名"}
SHAREHOLDER_KEYS = {"shareholder", "shareholder_name", "shareholderName", "shareholder_names", "shareholderNames", "shareholders", "股东", "股东姓名", "股东名称", "股东信息", "股东列表"}
RELATED_PARTY_KEYS = {
    "actual_controller",
    "actualController",
    "actual_controller_name",
    "actualControllerName",
    "controller",
    "controller_name",
    "controllerName",
    "beneficial_owner",
    "beneficialOwner",
    "beneficial_owner_name",
    "beneficialOwnerName",
    "ultimate_beneficial_owner",
    "ultimateBeneficialOwner",
    "ultimate_beneficial_owner_name",
    "ultimateBeneficialOwnerName",
    "spouse",
    "related_party",
    "relatedParty",
    "related_parties",
    "relatedParties",
    "related_party_names",
    "relatedPartyNames",
    "实控人",
    "实际控制人",
    "受益人",
    "最终受益人",
    "配偶",
    "关联方",
    "关联方名称",
    "关联方信息",
}
PARTY_NAME_KEYS = {"name", "company_name", "companyName", "person_name", "personName", "shareholder_name", "shareholderName", "姓名", "名称", "企业名称", "公司名称", "股东名称"}


def _add_excluded_party(excluded_parties: dict[str, str], value: Any, reason: str) -> None:
    name = _normalize_party_name(value)
    if name:
        excluded_parties.setdefault(name, reason)


def _collect_party_names(value: Any, reason: str, excluded_parties: dict[str, str]) -> None:
    if value is None:
        return
    if isinstance(value, (str, int, float, Decimal)):
        _add_excluded_party(excluded_parties, value, reason)
        return
    if isinstance(value, dict):
        matched = False
        for key, item in value.items():
            key_text = str(key)
            if key_text in PARTY_NAME_KEYS:
                _add_excluded_party(excluded_parties, item, reason)
                matched = True
            elif isinstance(item, (dict, list, tuple, set)):
                _collect_party_names(item, reason, excluded_parties)
        if not matched and len(value) == 1:
            _collect_party_names(next(iter(value.values())), reason, excluded_parties)
        return
    if isinstance(value, (list, tuple, set)):
        for item in value:
            _collect_party_names(item, reason, excluded_parties)


def _collect_excluded_parties_from_source(source: Any, excluded_parties: dict[str, str]) -> None:
    if not isinstance(source, dict):
        return
    for key, value in source.items():
        key_text = str(key)
        if key_text in SELF_PARTY_KEYS:
            _collect_party_names(value, "本方同名划转", excluded_parties)
        elif key_text in CUSTOMER_CONTAINER_KEYS:
            _collect_party_names(value, "本方同名划转", excluded_parties)
        elif key_text in LEGAL_REP_KEYS:
            _collect_party_names(value, "法定代表人往来", excluded_parties)
        elif key_text in SHAREHOLDER_KEYS:
            _collect_party_names(value, "股东往来", excluded_parties)
        elif key_text in RELATED_PARTY_KEYS:
            _collect_party_names(value, "关联方往来", excluded_parties)
        elif isinstance(value, dict):
            _collect_excluded_parties_from_source(value, excluded_parties)
        elif isinstance(value, (list, tuple, set)):
            for item in value:
                _collect_excluded_parties_from_source(item, excluded_parties)


def _is_placeholder(value: Any) -> bool:
    return str(value).strip() in PLACEHOLDER_VALUES


def _money(value: Any) -> Decimal | None:
    if value is None or _is_placeholder(value):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _text(value).replace(",", "").replace("￥", "").replace("¥", "").strip()
    if not text or text in {"-", "--"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return Decimal(match.group(0))
    except InvalidOperation:
        return None


def _int(value: Any) -> int | None:
    amount = _money(value)
    return int(amount) if amount is not None else None


def _fmt_money(value: Any) -> str:
    amount = _money(value)
    return f"{amount:,.2f}" if amount is not None else "0.00"


def _fmt_percent(numerator: Decimal, denominator: Decimal) -> str:
    if not denominator:
        return "0.00%"
    return f"{(numerator / denominator * Decimal('100')):.2f}%"


def _fmt_count(value: Any) -> str:
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return "0"


def _column_letter(index: int) -> str:
    if index <= 0:
        return ""
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _display(value: Any, default: str = UNKNOWN) -> str:
    text = _text(value)
    if not text or text in {"null", "None", "undefined", "{}", "[]"}:
        return default
    return text.replace("|", "｜")


def _date_text(value: Any, with_time: bool = False) -> str:
    if value is None or _is_placeholder(value):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S" if with_time else "%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = _text(value)
    if not text:
        return ""
    match = re.search(r"((?:19|20)\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})(?:日)?(?:\s+(\d{1,2}:\d{2}(?::\d{2})?))?", text)
    if match:
        base = f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
        time_part = match.group(4) or ""
        if with_time and time_part:
            if len(time_part.split(":")) == 2:
                time_part += ":00"
            return f"{base} {time_part}"
        return base
    match = re.search(r"\b((?:19|20)\d{2})(\d{2})(\d{2})\b", text)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return ""


def _date_range(value: Any) -> tuple[str, str]:
    text = _text(value)
    dates = re.findall(r"(?:19|20)\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}", text)
    if len(dates) >= 2:
        return _date_text(dates[0]), _date_text(dates[1])
    compact = re.findall(r"\b(?:19|20)\d{6}\b", text)
    if len(compact) >= 2:
        return _date_text(compact[0]), _date_text(compact[1])
    return "", ""


def _contains_any(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def _is_unknown_counterparty(value: Any) -> bool:
    text = _display(value, "")
    normalized = _normalize_party_name(text)
    normalized_unknowns = {_normalize_party_name(item) for item in UNKNOWN_COUNTERPARTY_NAMES}
    return not normalized or text in UNKNOWN_COUNTERPARTY_NAMES or normalized in normalized_unknowns


def _is_personal_name(value: Any) -> bool:
    text = _normalize_party_name(value)
    if not re.fullmatch(r"[\u4e00-\u9fff]{2,4}", text):
        return False
    return not any(keyword in text for keyword in ORGANIZATION_KEYWORDS)


def _is_noise_counterparty(value: Any, joined: str = "") -> bool:
    text = _display(value, "")
    if _is_unknown_counterparty(text):
        return True
    combined = f"{text} {joined}"
    return _contains_any(combined, NOISE_COUNTERPARTY_KEYWORDS)


def _is_organization_counterparty(value: Any) -> bool:
    text = _display(value, "")
    if _is_unknown_counterparty(text) or _is_personal_name(text):
        return False
    return bool(text) and not _is_noise_counterparty(text)


def _is_top_eligible_operating_tx(tx: dict[str, Any], direction: str) -> bool:
    if tx.get("direction") != direction:
        return False
    if direction == "in" and not tx.get("is_operating_inflow"):
        return False
    if direction == "out" and not tx.get("is_operating_outflow"):
        return False
    blocked_flags = (
        "is_excluded_related_party",
        "is_self_transfer",
        "is_related_party_transfer",
        "is_loan_related",
        "is_fee",
        "is_interest",
        "is_noise",
        "is_personal_counterparty",
    )
    if any(tx.get(flag) for flag in blocked_flags):
        return False
    if direction == "out" and any(tx.get(flag) for flag in ("is_salary", "is_tax")):
        return False
    return not _is_unknown_counterparty(tx.get("counterparty_name"))


def _operating_evidence(tx: dict[str, Any], direction: str) -> str:
    keywords = OPERATING_IN_KEYWORDS if direction == "in" else OPERATING_OUT_KEYWORDS
    candidates = [_text(tx.get(key)) for key in ("purpose", "summary", "remark")]
    for candidate in candidates:
        if candidate and _contains_any(candidate, keywords):
            return _normalize_evidence_label(candidate, keywords)
    joined = " ".join(candidates)
    for keyword in keywords:
        if keyword in joined:
            return keyword
    return ""


def _normalize_evidence_label(value: str, keywords: tuple[str, ...]) -> str:
    text = _text(value)
    if not text:
        return ""
    text = re.sub(r"20\d{2}年?", "", text)
    text = re.sub(r"[（(][^）)]{0,30}[）)]", "", text)
    text = re.sub(r"第?\d+批", "", text)
    text = re.sub(r"\d+(?:[./-]\d+)*", "", text)
    text = re.sub(r"[，,、；;:\s]+", "", text)
    for keyword in sorted(keywords, key=len, reverse=True):
        idx = text.find(keyword)
        if idx >= 0:
            start = max(0, idx - 8)
            end = min(len(text), idx + len(keyword) + 4)
            return text[start:end][:20]
    return text[:20]


def _bank_from_filename(filename: str) -> str:
    if "工商银行" in filename or "工行" in filename:
        return "工商银行"
    if "上海银行" in filename:
        return "上海银行"
    return ""


@dataclass
class AccountInfo:
    bank_name: str = ""
    account_name: str = ""
    account_no: str = ""
    branch_name: str = ""
    currency: str = "人民币"
    date_start: str = ""
    date_end: str = ""
    source_file: str = ""
    sheet_name: str = ""
    account_confidence: str = "medium"
    parse_status: str = "success"
    parse_warnings: list[str] = field(default_factory=list)


@dataclass
class FileParseResult:
    source_file: str
    sheet_name: str
    bank_name: str
    header_row_no: int
    header_col_start: int
    account: AccountInfo
    transactions: list[dict[str, Any]]
    raw_summary: dict[str, Any]
    placeholder_cleaned_count: int = 0
    status: str = "成功"
    warnings: list[str] = field(default_factory=list)


def _read_workbook(file_path: str, filename: str) -> list[tuple[str, list[list[Any]]]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows: list[list[Any]] = []
        raw = path.read_bytes()
        text = ""
        for encoding in ("utf-8-sig", "gb18030", "gbk"):
            try:
                text = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not text:
            text = raw.decode("utf-8", errors="ignore")
        for row in csv.reader(text.splitlines()):
            rows.append(row)
        return [(Path(filename).stem or "CSV", rows)]
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        last_error: Exception | None = None
        for read_only in (False, True):
            try:
                wb = load_workbook(path, read_only=read_only, data_only=True)
            except Exception as exc:
                last_error = exc
                logger.warning(
                    "[BankReconciliationDetail] workbook_open_failed filename=%s path=%s read_only=%s error=%s",
                    filename,
                    file_path,
                    read_only,
                    exc,
                )
                continue
            try:
                logger.info(
                    "[BankReconciliationDetail] workbook_opened filename=%s path=%s read_only=%s active_sheet=%s sheets=%s",
                    filename,
                    file_path,
                    read_only,
                    wb.active.title if wb.active else "",
                    wb.sheetnames,
                )
                sheets: list[tuple[str, list[list[Any]]]] = []
                for ws in wb.worksheets:
                    try:
                        dimension = ws.calculate_dimension()
                    except Exception as exc:
                        dimension = f"calculate_dimension_failed:{exc}"
                    logger.info(
                        "[BankReconciliationDetail] sheet_dimension file=%s sheet=%s read_only=%s max_row=%s max_col=%s dimension=%s",
                        file_path,
                        ws.title,
                        read_only,
                        getattr(ws, "max_row", ""),
                        getattr(ws, "max_column", ""),
                        dimension,
                    )
                    if read_only and dimension == "A1:A1" and hasattr(ws, "reset_dimensions"):
                        logger.warning(
                            "[BankReconciliationDetail] worksheet dimension is A1:A1, reset_dimensions file=%s sheet=%s",
                            file_path,
                            ws.title,
                        )
                        ws.reset_dimensions()
                        try:
                            dimension = ws.calculate_dimension(force=True)
                        except Exception as exc:
                            dimension = f"force_calculate_dimension_failed:{exc}"
                        logger.info(
                            "[BankReconciliationDetail] sheet_dimension_after_reset file=%s sheet=%s max_row=%s max_col=%s dimension=%s",
                            file_path,
                            ws.title,
                            getattr(ws, "max_row", ""),
                            getattr(ws, "max_column", ""),
                            dimension,
                        )
                    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
                    sheets.append((ws.title, rows))
                return sheets
            finally:
                wb.close()
        raise RuntimeError(f"workbook 读取失败：{last_error}") if last_error else RuntimeError("workbook 读取失败")
    if suffix == ".xls":
        try:
            import pandas as pd  # type: ignore

            sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
            return [(str(name), frame.where(frame.notna(), None).values.tolist()) for name, frame in sheets.items()]
        except Exception as exc:
            raise RuntimeError(f"暂不支持读取该 .xls 文件，请转换为 xlsx 后重试：{exc}") from exc
    raise RuntimeError(f"不支持的文件格式：{suffix or '未知'}")


PDF_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("交易时间", "交易日期", "记账日期", "入账日期", "发生日期", "日期"),
    "direction": ("交易方向", "借贷标志", "借/贷", "收支", "借方/贷方", "收入/支出"),
    "amount": ("交易金额", "发生额", "借方金额", "贷方金额", "转入金额", "转出金额", "收入金额", "支出金额", "收入", "支出"),
    "balance": ("余额", "账户余额"),
    "counterparty": ("对方户名", "对手名称", "对方单位", "对方名称", "对方账号", "对手账号"),
    "summary": ("摘要|备注", "摘要", "用途", "交易用途", "附言", "备注", "交易对手信息"),
}

PDF_BANK_NAMES = (
    "工商银行",
    "上海银行",
    "建设银行",
    "农业银行",
    "中国银行",
    "招商银行",
    "交通银行",
    "浦发银行",
    "民生银行",
    "平安银行",
    "网商银行",
    "泰隆银行",
    "齐鲁银行",
)


def _metadata_pages(metadata: dict[str, Any] | None) -> list[dict[str, Any]]:
    if not isinstance(metadata, dict):
        return []
    pages = metadata.get("raw_pages") or metadata.get("pages") or []
    return pages if isinstance(pages, list) else []


def _pdf_page_text(page: dict[str, Any]) -> str:
    parts = [str(page.get("text") or "")]
    for row in page.get("table_rows") or []:
        if isinstance(row, (list, tuple)):
            line = " ".join(_text(cell) for cell in row if _text(cell))
        else:
            line = _text(row)
        if line:
            parts.append(line)
    return "\n".join(part for part in parts if part)


def _load_pdf_pages(file_path: str, metadata: dict[str, Any] | None = None) -> tuple[list[dict[str, Any]], int, bool, int, str]:
    raw_pages = _metadata_pages(metadata)
    native_pages: list[dict[str, Any]] = []
    page_count = 0
    fail_reason = ""
    try:
        import fitz  # type: ignore

        with fitz.open(str(file_path)) as document:
            page_count = len(document)
            for page_no, page in enumerate(document, start=1):
                native_pages.append({"page": page_no, "text": page.get_text("text") or "", "source": "pdf_native"})
        try:
            import pdfplumber  # type: ignore

            with pdfplumber.open(str(file_path)) as pdf:
                for page_no, page in enumerate(pdf.pages, start=1):
                    if page_no > len(native_pages):
                        native_pages.append({"page": page_no, "text": page.extract_text() or "", "source": "pdfplumber"})
                    table_rows: list[list[str]] = []
                    for table in page.extract_tables() or []:
                        for row in table or []:
                            cells = [str(cell or "").strip() for cell in row]
                            if any(cells):
                                table_rows.append(cells)
                    if table_rows:
                        native_pages[page_no - 1].setdefault("table_rows", [])
                        native_pages[page_no - 1]["table_rows"].extend(table_rows)
        except Exception as exc:  # pragma: no cover - pdfplumber is an optional enhancement
            logger.info("[BankReconciliationDetail][PDF] pdfplumber_table_extract_skipped file=%s error=%s", file_path, exc)
    except Exception as exc:
        fail_reason = f"PDF 文件读取失败：{exc}"
        if raw_pages:
            logger.warning("[BankReconciliationDetail][PDF] native_read_failed_use_raw_pages file=%s error=%s", file_path, exc)
        else:
            logger.exception("[BankReconciliationDetail][PDF] read_failed file=%s", file_path)

    native_text_length = sum(len(_pdf_page_text(page)) for page in native_pages)
    raw_text_length = sum(len(_pdf_page_text(page)) for page in raw_pages)
    if native_text_length >= 20:
        return native_pages, page_count or len(native_pages), False, native_text_length, fail_reason
    if raw_text_length >= 20:
        logger.info(
            "[BankReconciliationDetail][PDF] native_text_short_use_raw_pages file=%s native_len=%s raw_len=%s",
            file_path,
            native_text_length,
            raw_text_length,
        )
        return raw_pages, page_count or len(raw_pages), True, raw_text_length, fail_reason
    if fail_reason:
        return [], page_count, False, native_text_length, fail_reason
    return native_pages, page_count or len(native_pages), False, native_text_length, "PDF 未提取到文本，OCR 也未识别到有效内容"


def _detect_pdf_bank_name(text: str, source_file: str) -> str:
    bank = _bank_from_filename(source_file)
    if bank:
        return bank
    for name in PDF_BANK_NAMES:
        if name in text or name in source_file:
            return name
    return UNKNOWN


def _pdf_value(text: str, labels: tuple[str, ...]) -> str:
    for label in labels:
        pattern = re.compile(rf"{re.escape(label)}\s*[:：]?\s*([^\n\r]{{1,100}})")
        match = pattern.search(text)
        if not match:
            continue
        value = match.group(1).strip()
        value = re.split(r"\s+(?:户名|账户名称|客户名称|账号|账户|开户行|币种|银行名称)\s*[:：]?", value, maxsplit=1)[0]
        return _text(value)
    return ""


def _parse_pdf_account(text: str, source_file: str, sheet_name: str, transactions: list[dict[str, Any]] | None = None) -> AccountInfo:
    account = AccountInfo(bank_name=_detect_pdf_bank_name(text, source_file), source_file=source_file, sheet_name=sheet_name)
    account.account_name = _pdf_value(text, ("户名", "账户名称", "客户名称"))
    account.account_no = re.sub(r"\D", "", _pdf_value(text, ("账号", "账户号码", "账户", "银行账号"))) or ""
    account.branch_name = _pdf_value(text, ("开户行", "开户网点"))
    account.currency = _pdf_value(text, ("币种",)) or "人民币"
    start, end = _date_range(text)
    account.date_start = start
    account.date_end = end
    if transactions:
        dates = sorted(tx.get("accounting_date") for tx in transactions if tx.get("accounting_date"))
        if dates:
            account.date_start = account.date_start or dates[0]
            account.date_end = account.date_end or dates[-1]
    return account


def _pdf_header_hit_count(line: str) -> int:
    compact = _compact(line)
    hits = 0
    for aliases in PDF_HEADER_ALIASES.values():
        if any(_compact(alias) in compact for alias in aliases):
            hits += 1
    return hits


def _detect_pdf_header(pages: list[dict[str, Any]]) -> tuple[bool, int, int, str]:
    for page in pages:
        page_no = int(page.get("page") or 0)
        for line_no, line in enumerate(_pdf_page_text(page).splitlines(), start=1):
            if _pdf_header_hit_count(line) >= 4:
                return True, page_no, line_no, line
    return False, 0, 0, ""


def _is_qilu_pdf_detail(text: str, source_file: str) -> bool:
    source = f"{source_file}\n{text}"
    if "齐鲁银行" in source:
        return True
    if "单位活期存款账户交易明细" in source:
        return True
    compact = _compact(source)
    required = ("记账日期", "交易渠道", "收入", "支出", "账户余额", "摘要|备注", "交易对手信息")
    return sum(1 for item in required if _compact(item) in compact) >= 5


def _capture_pdf_label(text: str, labels: tuple[str, ...], *, max_len: int = 120) -> str:
    stop = r"(?:账号|账户名称|开户机构|开户行|起止日期|交易方向|币种|收入金额合计|支出金额合计|第\d+/\d+页|共\d+条)"
    for label in labels:
        pattern = re.compile(rf"{re.escape(label)}\s*[:：]?\s*([^\n\r]{{1,{max_len}}})")
        match = pattern.search(text)
        if not match:
            continue
        value = re.split(rf"\s+{stop}\s*[:：]?", match.group(1), maxsplit=1)[0]
        value = re.split(stop, value, maxsplit=1)[0]
        return _text(value)
    return ""


def _qilu_pdf_account_and_summary(text: str, source_file: str, sheet_name: str) -> tuple[AccountInfo, dict[str, Any]]:
    account = AccountInfo(bank_name="齐鲁银行", source_file=source_file, sheet_name=sheet_name)
    account.account_no = re.sub(r"\D", "", _capture_pdf_label(text, ("账号",), max_len=40))
    account.account_name = _capture_pdf_label(text, ("账户名称", "户名"), max_len=120)
    account.branch_name = _capture_pdf_label(text, ("开户机构", "开户行"), max_len=120)
    account.currency = _capture_pdf_label(text, ("币种",), max_len=20) or "人民币"
    header_lines = normalize_qilu_pdf_text(text).splitlines()[:80]
    for idx in range(max(0, len(header_lines) - 5)):
        current = _compact(header_lines[idx]).rstrip(":：")
        second = _compact(header_lines[idx + 1]).rstrip(":：")
        third = _compact(header_lines[idx + 2]).rstrip(":：")
        if current == _compact("账号") and second == _compact("账户名称") and third == _compact("起止日期"):
            values = [_text(line) for line in header_lines[idx + 3 : idx + 12] if _text(line)]
            logger.info("qilu header lines around account labels=%s", header_lines[idx : idx + 8])
            if len(values) >= 3:
                account_no, account_name, date_range = values[0], values[1], values[2]
                if re.fullmatch(r"\d{8,30}", account_no):
                    account.account_no = account_no
                if any(keyword in account_name for keyword in ("公司", "有限公司", "工程", "建设", "商行", "个体工商户", "合作社")):
                    account.account_name = account_name
                start, end = _date_range(date_range)
                if start and end:
                    account.date_start = start
                    account.date_end = end
                logger.info(
                    "qilu account_no=%s account_name=%s date_range=%s",
                    account.account_no,
                    account.account_name,
                    date_range,
                )
            break
    date_match = re.search(
        r"起止日期\s*[:：]?\s*((?:19|20)\d{2}[-/.]\d{1,2}[-/.]\d{1,2})\s*[-—至]\s*((?:19|20)\d{2}[-/.]\d{1,2}[-/.]\d{1,2})",
        text,
    )
    if date_match and not (account.date_start and account.date_end):
        account.date_start = _date_text(date_match.group(1))
        account.date_end = _date_text(date_match.group(2))
    if not account.account_no:
        logger.info("qilu account_no not recognized header_first_lines=%s", header_lines[:30])
    raw_summary: dict[str, Any] = {}
    count_match = re.search(r"共\s*(\d+)\s*条", text)
    if count_match:
        raw_summary["raw_transaction_count"] = _int(count_match.group(1))
    for label, key in (("收入金额合计", "income_total"), ("支出金额合计", "out_total")):
        match = re.search(rf"{label}\s*[:：]?\s*(-?(?:\d{{1,3}}(?:,\d{{3}})+|\d+)(?:\.\d+)?)", text)
        if match:
            raw_summary[key] = _money(match.group(1))
    return account, raw_summary


def _qilu_counterparty_from_text(text: str) -> dict[str, str]:
    result = {"counterparty_account": "", "counterparty_name": "", "counterparty_bank_no": ""}
    source = _text(text)
    account_match = re.search(r"(?:对方账号|交易对手账号|账号)\s*[:：]?\s*([0-9]{6,32})", source)
    if account_match:
        result["counterparty_account"] = account_match.group(1)
    name_match = re.search(r"(?:对方户名|对方名称|交易对手名称|户名)\s*[:：]?\s*([\u4e00-\u9fffA-Za-z0-9（）()·]{2,80})", source)
    if name_match:
        result["counterparty_name"] = _text(name_match.group(1))
    bank_match = re.search(r"(?:对方开户行|对方行名|对方开户机构|清算行|开户行)\s*[:：]?\s*([\u4e00-\u9fffA-Za-z0-9（）()·]{2,120})", source)
    if bank_match:
        result["counterparty_bank_no"] = _text(bank_match.group(1))
    if not result["counterparty_name"]:
        org_match = re.search(r"([\u4e00-\u9fffA-Za-z0-9（）()·]{2,80}(?:有限公司|有限责任公司|银行|公司|中心|集团|商行|合作社|个体工商户|厂|店|局|院|所))", source)
        if org_match:
            result["counterparty_name"] = _text(org_match.group(1))
    return result


def _qilu_make_tx(
    *,
    source_file: str,
    account: AccountInfo,
    row_no: int,
    accounting_date: str,
    channel: str,
    income: Decimal | None,
    out_amount: Decimal | None,
    balance: Decimal | None,
    summary: str,
    counterparty: dict[str, str] | None = None,
) -> dict[str, Any] | None:
    income = income or Decimal("0")
    out_amount = out_amount or Decimal("0")
    if income > 0 and out_amount == 0:
        direction = "in"
        amount = income
    elif out_amount > 0 and income == 0:
        direction = "out"
        amount = out_amount
    else:
        return None
    counterparty = counterparty or {}
    tx = {
        "transaction_id": "",
        "source_file": source_file,
        "bank_name": "齐鲁银行",
        "account_name": account.account_name,
        "account_no": account.account_no,
        "trade_time": accounting_date,
        "accounting_date": accounting_date,
        "direction": direction,
        "direction_name": "入账" if direction == "in" else "出账",
        "amount": amount,
        "balance": balance,
        "counterparty_name": counterparty.get("counterparty_name") or "",
        "counterparty_account": counterparty.get("counterparty_account") or "",
        "counterparty_bank_no": counterparty.get("counterparty_bank_no") or "",
        "summary": _text(summary),
        "purpose": _text(summary),
        "remark": _text(channel),
        "voucher_no": "",
        "raw_row_no": row_no,
        "warning": "",
    }
    _classify(tx, {})
    return tx


def _qilu_parse_table_row(row: list[Any], source_file: str, account: AccountInfo, row_no: int) -> dict[str, Any] | None:
    cells = [_text(cell) for cell in row]
    if not any(cells):
        return None
    date_idx = next((idx for idx, cell in enumerate(cells) if _date_text(cell)), -1)
    if date_idx < 0:
        return None
    joined_cells = " ".join(cells)
    if any(marker in joined_cells for marker in ("起止日期", "收入金额合计", "支出金额合计")):
        return None
    if re.search(r"第\s*\d+\s*/\s*\d+\s*页|共\s*\d+\s*条", joined_cells):
        return None
    accounting_date = _date_text(cells[date_idx])
    channel = cells[date_idx + 1] if date_idx + 1 < len(cells) else ""
    income = _money(cells[date_idx + 2]) if date_idx + 2 < len(cells) else None
    out_amount = _money(cells[date_idx + 3]) if date_idx + 3 < len(cells) else None
    balance = _money(cells[date_idx + 4]) if date_idx + 4 < len(cells) else None
    summary = " ".join(cell for cell in cells[date_idx + 5 :] if cell)
    counterparty = _qilu_counterparty_from_text(" ".join(cells))
    return _qilu_make_tx(
        source_file=source_file,
        account=account,
        row_no=row_no,
        accounting_date=accounting_date,
        channel=channel,
        income=income,
        out_amount=out_amount,
        balance=balance,
        summary=summary,
        counterparty=counterparty,
    )


def _qilu_parse_text_line(line: str, source_file: str, account: AccountInfo, row_no: int) -> dict[str, Any] | None:
    text = _text(line)
    if not text or any(marker in text for marker in ("单位活期存款账户交易明细", "起止日期", "收入金额合计", "支出金额合计")):
        return None
    if re.search(r"第\s*\d+\s*/\s*\d+\s*页|共\s*\d+\s*条", text):
        return None
    accounting_date = _date_text(text)
    if not accounting_date:
        return None
    money_values = [_money(item) for item in re.findall(r"-?(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}", text)]
    money_values = [item for item in money_values if item is not None]
    if len(money_values) < 3:
        return None
    income, out_amount, balance = money_values[0], money_values[1], money_values[2]
    cleaned = _strip_pdf_noise(text)
    counterparty = _qilu_counterparty_from_text(cleaned)
    summary = cleaned
    return _qilu_make_tx(
        source_file=source_file,
        account=account,
        row_no=row_no,
        accounting_date=accounting_date,
        channel="",
        income=income,
        out_amount=out_amount,
        balance=balance,
        summary=summary,
        counterparty=counterparty,
    )


QILU_AMOUNT_PATTERN = re.compile(r"[-+]?\d{1,3}(?:,\d{3})*(?:\.\d{2})|[-+]?\d+(?:\.\d{2})")
QILU_BLOCK_AMOUNT = r"(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}"
QILU_BANK_START_KEYWORDS = (
    "中国工商银行",
    "中国农业银行",
    "中国银行",
    "中国建设银行",
    "交通银行",
    "招商银行",
    "中信银行",
    "上海浦东发展银行",
    "浦发银行",
    "齐鲁银行",
    "德州银行",
    "郑州银行",
    "北京银行",
    "济宁银行",
    "山东省农村信用社",
    "天津农行支付系统处理中心",
    "中国工商银行总行清算中心",
    "中国农业银行资金清算中心",
    "中国建设银行总行",
)


def _qilu_amounts(line: str) -> list[Decimal]:
    amounts: list[Decimal] = []
    for item in QILU_AMOUNT_PATTERN.findall(_text(line)):
        amount = _money(item)
        if amount is not None:
            amounts.append(amount)
    return amounts


def normalize_qilu_pdf_text(text: str) -> str:
    lines: list[str] = []
    for raw in str(text or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = _text(raw)
        if not line:
            continue
        if re.fullmatch(r"(?:19|20)\d{2}[./-]\d{1,2}[./-]\d{1,2}\s+\d{1,2}:\d{2}:\d{2}", line):
            continue
        line = line.replace("交易对手信息:", "交易对手信息：")
        line = re.sub(r"\s+", " ", line).strip()
        lines.append(line)
    return "\n".join(lines)


def _qilu_clean_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw in normalize_qilu_pdf_text(text).splitlines():
        line = _text(raw)
        if not line:
            continue
        if re.fullmatch(r"(?:19|20)\d{2}[./-]\d{1,2}[./-]\d{1,2}\s+\d{1,2}:\d{2}:\d{2}", line):
            continue
        if re.search(r"第\s*\d+\s*/\s*\d+\s*页|共\s*\d+\s*条", line):
            continue
        if any(marker in line for marker in ("单位活期存款账户交易明细", "开户机构", "账户名称", "起止日期", "收入金额合计", "支出金额合计", "交易方向：全部", "币种：")):
            continue
        lines.append(line)
    return lines


def _qilu_is_counterparty_line(line: str) -> bool:
    text = _text(line)
    return bool(re.match(r"^[0-9A-Za-z]{6,}\s+", text))


def _qilu_split_counterparty_line(cp_line: str) -> tuple[str, str]:
    text = _text(cp_line)
    if not text:
        return "", ""
    best_idx = -1
    for keyword in QILU_BANK_START_KEYWORDS:
        idx = text.find(keyword)
        if idx >= 0 and (best_idx < 0 or idx < best_idx):
            best_idx = idx
    if best_idx > 0:
        return _text(text[:best_idx]), _text(text[best_idx:])
    bank_keywords = ("银行", "信用社", "清算中心", "支行", "分行", "总行", "农村商业银行", "农商行", "支付系统处理中心")
    for keyword in bank_keywords:
        idx = text.find(keyword)
        if idx > 0:
            space_idx = text.rfind(" ", 0, idx)
            if space_idx > 0:
                return _text(text[:space_idx]), _text(text[space_idx + 1 :])
    return text, ""


def _qilu_parse_counterparty_line(line: str) -> dict[str, str]:
    result = {"counterparty_account": "", "counterparty_name": "", "counterparty_bank_no": ""}
    text = _text(line)
    match = re.match(r"^([0-9A-Za-z]{6,})\s+(.+)$", text)
    if not match:
        return _qilu_counterparty_from_text(text)
    result["counterparty_account"] = match.group(1)
    rest = _text(match.group(2))
    parts = [part for part in re.split(r"\s+", rest) if part]
    if len(parts) >= 2:
        result["counterparty_name"] = parts[0]
        result["counterparty_bank_no"] = " ".join(parts[1:])
        return result
    bank_keywords = ("银行", "信用社", "清算中心", "支行", "分行", "总行", "农商行", "农村商业银行")
    bank_start = -1
    for keyword in bank_keywords:
        idx = rest.find(keyword)
        if idx >= 0 and (bank_start < 0 or idx < bank_start):
            bank_start = idx
    if bank_start > 0:
        company_end = rest.rfind(" ", 0, bank_start)
        if company_end > 0:
            result["counterparty_name"] = _text(rest[:company_end])
            result["counterparty_bank_no"] = _text(rest[company_end + 1 :])
        else:
            org_match = re.match(r"(.+?(?:有限公司|有限责任公司|公司|中心|集团|商行|合作社|个体工商户|厂|店|局|院|所))(.+)$", rest)
            if org_match:
                result["counterparty_name"] = _text(org_match.group(1))
                result["counterparty_bank_no"] = _text(org_match.group(2))
    if not result["counterparty_name"]:
        result["counterparty_name"] = rest
    return result


def _qilu_date_summary(line: str) -> tuple[str, str, str]:
    text = _text(line)
    match = re.match(r"^((?:19|20)\d{2}[-/.]\d{1,2}[-/.]\d{1,2})\s+(\S+)\s+(.+)$", text)
    if not match:
        return "", "", ""
    return _date_text(match.group(1)), _text(match.group(2)), _text(match.group(3))


def parse_qilu_transactions_by_regex(full_text: str, source_file: str, account: AccountInfo) -> list[dict[str, Any]]:
    normalized = normalize_qilu_pdf_text(full_text)
    lines = normalized.splitlines()
    pattern = re.compile(
        rf"(?P<income>{QILU_BLOCK_AMOUNT})\s+(?P<out>{QILU_BLOCK_AMOUNT})\s*\n"
        rf"(?P<cp_account>[0-9A-Za-z]{{6,}})\s+(?P<cp_line>.+?)\n"
        rf"(?P<balance>{QILU_BLOCK_AMOUNT})\s*\n"
        rf"交易对手信息[:：]\s*\n"
        rf"(?P<seq>\d{{1,5}})\s*\n"
        rf"(?P<date>20\d{{2}}-\d{{2}}-\d{{2}})\s+(?P<channel>\S+)\s+(?P<summary>[^\n]+)",
        re.MULTILINE,
    )
    transactions: list[dict[str, Any]] = []
    for match in pattern.finditer(normalized):
        income = _money(match.group("income")) or Decimal("0")
        out_amount = _money(match.group("out")) or Decimal("0")
        if income == 0 and out_amount == 0:
            continue
        counterparty_name, counterparty_bank_name = _qilu_split_counterparty_line(match.group("cp_line"))
        tx = _qilu_make_tx(
            source_file=source_file,
            account=account,
            row_no=_int(match.group("seq")) or len(transactions) + 1,
            accounting_date=_date_text(match.group("date")),
            channel=match.group("channel"),
            income=income,
            out_amount=out_amount,
            balance=_money(match.group("balance")),
            summary=match.group("summary"),
            counterparty={
                "counterparty_account": match.group("cp_account"),
                "counterparty_name": counterparty_name,
                "counterparty_bank_no": counterparty_bank_name,
            },
        )
        if tx:
            transactions.append(tx)
    marker_count = normalized.count("交易对手信息")
    date_line_count = sum(1 for line in lines if re.match(r"^20\d{2}-\d{2}-\d{2}\s+\S+\s+", line))
    amount_pair_line_count = sum(1 for line in lines if len(_qilu_amounts(line)) == 2)
    logger.info(
        "qilu regex full_text_length=%s marker_count=%s date_line_count=%s amount_pair_line_count=%s parsed_transaction_count=%s",
        len(normalized),
        marker_count,
        date_line_count,
        amount_pair_line_count,
        len(transactions),
    )
    if not transactions:
        logger.warning("qilu regex parsed 0 transactions, first_lines=%s", "\n".join(lines[:120]))
    return transactions


def _qilu_parse_transaction_blocks_from_text(text: str, source_file: str, account: AccountInfo) -> list[dict[str, Any]]:
    lines = _qilu_clean_lines(text)
    marker_indexes = [idx for idx, line in enumerate(lines) if "交易对手信息" in line]
    logger.info("qilu pdf text lines=%s", len(lines))
    logger.info("qilu markers transaction_counterparty_info=%s", len(marker_indexes))
    transactions: list[dict[str, Any]] = []
    skipped_debug: list[tuple[int, str, list[str]]] = []
    for marker_idx in marker_indexes:
        before_start = max(0, marker_idx - 6)
        before = lines[before_start:marker_idx]
        after = lines[marker_idx + 1 : marker_idx + 7]
        debug_block = before + [lines[marker_idx]] + after
        balance: Decimal | None = None
        counterparty: dict[str, str] = {}
        income: Decimal | None = None
        out_amount: Decimal | None = None
        for line in reversed(before[-3:]):
            amounts = _qilu_amounts(line)
            if len(amounts) == 1:
                balance = amounts[0]
                break
        counterparty_idx = -1
        for local_idx, line in enumerate(before):
            if _qilu_is_counterparty_line(line):
                counterparty_idx = local_idx
                counterparty = _qilu_parse_counterparty_line(line)
        amount_search = before[:counterparty_idx] if counterparty_idx > 0 else before
        for line in reversed(amount_search):
            amounts = _qilu_amounts(line)
            if len(amounts) == 2:
                income, out_amount = amounts[0], amounts[1]
                break
        serial_no = ""
        accounting_date = ""
        channel = ""
        summary = ""
        for line in after[:3]:
            if re.fullmatch(r"\d{1,4}", line):
                serial_no = line
                break
        for line in after:
            accounting_date, channel, summary = _qilu_date_summary(line)
            if accounting_date:
                break
        reason = ""
        if income is None or out_amount is None:
            reason = "missing income/out amount line"
        elif not accounting_date:
            reason = "missing accounting date line"
        elif (income <= 0 and out_amount <= 0) or (income > 0 and out_amount > 0):
            reason = "invalid income/out amount"
        if reason:
            skipped_debug.append((marker_idx, reason, debug_block))
            logger.warning("qilu skipped block at marker=%s reason=%s block=%s", marker_idx, reason, debug_block)
            continue
        tx = _qilu_make_tx(
            source_file=source_file,
            account=account,
            row_no=_int(serial_no) or marker_idx + 1,
            accounting_date=accounting_date,
            channel=channel,
            income=income,
            out_amount=out_amount,
            balance=balance,
            summary=summary,
            counterparty=counterparty,
        )
        if tx and (tx.get("counterparty_name") or tx.get("summary")) and (tx.get("balance") is not None or tx.get("counterparty_account")):
            transactions.append(tx)
        else:
            skipped_debug.append((marker_idx, "incomplete transaction block", debug_block))
            logger.warning("qilu skipped block at marker=%s reason=%s block=%s", marker_idx, "incomplete transaction block", debug_block)
    income_sum = _sum(transactions, lambda tx: tx.get("direction") == "in")
    out_sum = _sum(transactions, lambda tx: tx.get("direction") == "out")
    logger.info("qilu parsed transaction count=%s", len(transactions))
    logger.info("qilu parsed income total=%s out total=%s", income_sum, out_sum)
    if not transactions and skipped_debug:
        logger.warning("qilu first skipped blocks=%s", skipped_debug[:3])
    return transactions


def _merge_qilu_counterparty(tx: dict[str, Any], text: str) -> None:
    counterparty = _qilu_counterparty_from_text(text)
    for key, value in counterparty.items():
        if value and not tx.get(key):
            tx[key] = value
    _classify(tx, {})


def _parse_qilu_pdf_bank_reconciliation_detail(
    pages: list[dict[str, Any]],
    text: str,
    source_file: str,
    page_count: int,
    header_detected: bool,
    header_line_no: int,
) -> tuple[FileParseResult | None, list[str]]:
    warnings: list[str] = []
    account, raw_summary = _qilu_pdf_account_and_summary(text, source_file, "PDF")
    raw_page_text = "\n".join(str(page.get("text") or "") for page in pages)
    transactions: list[dict[str, Any]] = parse_qilu_transactions_by_regex(raw_page_text or text, source_file, account)
    if not transactions:
        transactions = _qilu_parse_transaction_blocks_from_text(raw_page_text or text, source_file, account)
    last_tx: dict[str, Any] | None = None
    row_no = 0
    if not transactions:
        for page in pages:
            for row in page.get("table_rows") or []:
                if not isinstance(row, (list, tuple)):
                    continue
                row_no += 1
                tx = _qilu_parse_table_row(list(row), source_file, account, row_no)
                if tx:
                    transactions.append(tx)
                    last_tx = tx
                elif last_tx and "交易对手信息" in " ".join(_text(cell) for cell in row):
                    _merge_qilu_counterparty(last_tx, " ".join(_text(cell) for cell in row))
            for line in str(page.get("text") or "").splitlines():
                row_no += 1
                if "交易对手信息" in line and last_tx:
                    _merge_qilu_counterparty(last_tx, line)
                    continue
                tx = _qilu_parse_text_line(line, source_file, account, row_no)
                if tx:
                    signature = (tx.get("accounting_date"), tx.get("direction"), _fmt_money(tx.get("amount")), _fmt_money(tx.get("balance")), tx.get("summary"))
                    if not any((old.get("accounting_date"), old.get("direction"), _fmt_money(old.get("amount")), _fmt_money(old.get("balance")), old.get("summary")) == signature for old in transactions):
                        transactions.append(tx)
                        last_tx = tx
    if transactions:
        dates = sorted(tx.get("accounting_date") for tx in transactions if tx.get("accounting_date"))
        if dates:
            account.date_start = account.date_start or dates[0]
            account.date_end = account.date_end or dates[-1]
    if not raw_summary.get("raw_transaction_count"):
        raw_summary["raw_transaction_count"] = len(transactions)
    income_header = _money(raw_summary.get("income_total"))
    out_header = _money(raw_summary.get("out_total"))
    income_sum = _sum(transactions, lambda tx: tx.get("direction") == "in")
    out_sum = _sum(transactions, lambda tx: tx.get("direction") == "out")
    reconcile_ok = True
    if income_header is not None and abs(income_sum - income_header) > Decimal("0.01"):
        reconcile_ok = False
        warnings.append(f"收入金额合计与交易明细合计差异：页眉 {_fmt_money(income_header)}，明细 {_fmt_money(income_sum)}")
    if out_header is not None and abs(out_sum - out_header) > Decimal("0.01"):
        reconcile_ok = False
        warnings.append(f"支出金额合计与交易明细合计差异：页眉 {_fmt_money(out_header)}，明细 {_fmt_money(out_sum)}")
    expected_count = _int(raw_summary.get("raw_transaction_count"))
    if expected_count and expected_count != len(transactions):
        reconcile_ok = False
        warnings.append(f"交易笔数与页眉不一致：页眉 {expected_count} 条，明细 {len(transactions)} 条")
    logger.info(
        "[BankReconciliationDetail][PDF][Qilu] file_name=%s page_count=%s detected_bank_name=%s detected_pdf_format=qilu_unit_current_account_detail account_no=%s account_name=%s branch_name=%s date_start=%s date_end=%s header_detected=%s parsed_transaction_count=%s expected_transaction_count=%s income_total_from_header=%s out_total_from_header=%s income_total_from_transactions=%s out_total_from_transactions=%s amount_reconcile_status=%s",
        source_file,
        page_count,
        "齐鲁银行",
        account.account_no,
        account.account_name,
        account.branch_name,
        account.date_start,
        account.date_end,
        header_detected,
        len(transactions),
        expected_count or "",
        _fmt_money(income_header),
        _fmt_money(out_header),
        _fmt_money(income_sum),
        _fmt_money(out_sum),
        "ok" if reconcile_ok else "mismatch",
    )
    if not transactions:
        return None, [f"{source_file} 找到齐鲁银行交易明细格式，但未读取到交易行"]
    return FileParseResult(
        source_file=source_file,
        sheet_name="PDF",
        bank_name="齐鲁银行",
        header_row_no=header_line_no if header_detected else 0,
        header_col_start=1 if header_detected else 0,
        account=account,
        transactions=transactions,
        raw_summary=raw_summary,
        placeholder_cleaned_count=0,
        status="成功" if reconcile_ok else "部分成功",
        warnings=warnings,
    ), warnings


def _strip_pdf_noise(text: str) -> str:
    text = re.sub(r"(?:19|20)\d{2}[-/.年]\d{1,2}[-/.月]\d{1,2}(?:日)?(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?", " ", text)
    text = re.sub(r"-?(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}", " ", text)
    text = re.sub(r"\b\d{8,30}\b", " ", text)
    for token in (
        "交易时间",
        "交易日期",
        "记账日期",
        "交易方向",
        "借贷标志",
        "交易金额",
        "发生额",
        "余额",
        "账户余额",
        "对方户名",
        "对手名称",
        "对方单位",
        "对方名称",
        "对方账号",
        "对手账号",
        "摘要",
        "用途",
        "交易用途",
        "附言",
        "备注",
        "入账",
        "出账",
        "借方",
        "贷方",
        "收入",
        "支出",
        "借",
        "贷",
    ):
        text = text.replace(token, " ")
    return re.sub(r"\s+", " ", text).strip()


def _extract_counterparty_from_pdf_line(line: str) -> tuple[str, str]:
    cleaned = _strip_pdf_noise(line)
    org_match = re.search(r"([\u4e00-\u9fffA-Za-z0-9（）()·]{2,80}(?:有限公司|有限责任公司|银行|公司|中心|集团|商行|合作社|个体工商户|厂|店|局|院|所))", cleaned)
    if org_match:
        name = _text(org_match.group(1))
        return name, _text(cleaned.replace(name, " ", 1))
    person_match = re.search(r"(?<![\u4e00-\u9fff])([\u4e00-\u9fff]{2,4})(?![\u4e00-\u9fff])", cleaned)
    if person_match:
        name = _text(person_match.group(1))
        return name, _text(cleaned.replace(name, " ", 1))
    return "", _text(cleaned)


def _parse_pdf_transaction_line(line: str, source_file: str, bank_name: str, account: AccountInfo, row_no: int) -> dict[str, Any] | None:
    text = _text(line)
    if not text or _pdf_header_hit_count(text) >= 4:
        return None
    trade_time = _date_text(text, with_time=True)
    accounting_date = _date_text(text)
    if not trade_time and not accounting_date:
        return None
    amount_matches = re.findall(r"-?(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}", text)
    if not amount_matches:
        return None
    direction = ""
    if any(word in text for word in ("入账", "贷方", "收入", "转入")):
        direction = "in"
    if any(word in text for word in ("出账", "借方", "支出", "转出")):
        direction = "out"
    if not direction:
        if re.search(r"(?<![\u4e00-\u9fff])贷(?![\u4e00-\u9fff])", text):
            direction = "in"
        elif re.search(r"(?<![\u4e00-\u9fff])借(?![\u4e00-\u9fff])", text):
            direction = "out"
    if not direction:
        return None
    amount = _money(amount_matches[0])
    if amount is None:
        return None
    balance = _money(amount_matches[-1]) if len(amount_matches) >= 2 else None
    counterparty_name, remainder = _extract_counterparty_from_pdf_line(text)
    evidence_text = _text(remainder)
    tx = {
        "transaction_id": "",
        "source_file": source_file,
        "bank_name": bank_name,
        "account_name": account.account_name,
        "account_no": account.account_no,
        "trade_time": trade_time or accounting_date,
        "accounting_date": accounting_date or (trade_time[:10] if trade_time else ""),
        "direction": direction,
        "direction_name": "入账" if direction == "in" else "出账",
        "amount": amount,
        "balance": balance,
        "counterparty_name": counterparty_name,
        "counterparty_account": "",
        "counterparty_bank_no": "",
        "summary": evidence_text,
        "purpose": evidence_text,
        "remark": "",
        "voucher_no": "",
        "raw_row_no": row_no,
        "warning": "",
    }
    return tx


def parse_pdf_bank_reconciliation_detail(file_path: str, source_file: str, metadata: dict[str, Any] | None = None) -> tuple[FileParseResult | None, list[str]]:
    warnings: list[str] = []
    pages, page_count, used_ocr, text_length, load_reason = _load_pdf_pages(file_path, metadata)
    text = "\n".join(_pdf_page_text(page) for page in pages)
    bank_name = _detect_pdf_bank_name(text, source_file)
    header_detected, header_page_no, header_line_no, header_line = _detect_pdf_header(pages)
    logger.info(
        "[BankReconciliationDetail][PDF] file_name=%s file_ext=.pdf pdf_page_count=%s text_extract_length=%s used_ocr=%s detected_bank_name=%s header_detected=%s header_page_no=%s header_line=%s",
        source_file,
        page_count,
        text_length,
        used_ocr,
        bank_name,
        header_detected,
        header_page_no,
        header_line,
    )
    if not pages or text_length < 20:
        reason = load_reason or "PDF 未提取到文本，OCR 也未识别到有效内容"
        logger.error("[BankReconciliationDetail][PDF] parse_failed file=%s reason=%s", source_file, reason)
        return None, [f"{source_file} {reason}"]

    if _is_qilu_pdf_detail(text, source_file):
        parsed_qilu, qilu_warnings = _parse_qilu_pdf_bank_reconciliation_detail(
            pages,
            text,
            source_file,
            page_count,
            header_detected,
            header_line_no,
        )
        if parsed_qilu:
            return parsed_qilu, qilu_warnings
        return None, qilu_warnings

    account = _parse_pdf_account(text, source_file, "PDF")
    transactions: list[dict[str, Any]] = []
    row_no = 0
    for page in pages:
        for line in _pdf_page_text(page).splitlines():
            row_no += 1
            tx = _parse_pdf_transaction_line(line, source_file, bank_name, account, row_no)
            if not tx:
                continue
            transactions.append(tx)

    if transactions:
        account = _parse_pdf_account(text, source_file, "PDF", transactions)
        for tx in transactions:
            tx["account_name"] = account.account_name
            tx["account_no"] = account.account_no
            _classify(tx, {})
    if not transactions:
        reason = "PDF 中未找到交易明细表头" if not header_detected else "找到表头但未读取到交易行"
        logger.error("[BankReconciliationDetail][PDF] parse_failed file=%s reason=%s", source_file, reason)
        return None, [f"{source_file} {reason}"]

    if not any(tx.get("amount") is not None for tx in transactions):
        reason = "金额字段无法识别"
        logger.error("[BankReconciliationDetail][PDF] parse_failed file=%s reason=%s", source_file, reason)
        return None, [f"{source_file} {reason}"]
    if not any(tx.get("accounting_date") for tx in transactions):
        reason = "日期字段无法识别"
        logger.error("[BankReconciliationDetail][PDF] parse_failed file=%s reason=%s", source_file, reason)
        return None, [f"{source_file} {reason}"]

    raw_summary = {"raw_transaction_count": len(transactions)}
    if used_ocr:
        warnings.append("PDF 原生文本较少，已使用上传链路 OCR/页面文本")
    if not header_detected:
        warnings.append("PDF 未稳定识别交易表头，已按日期和金额行尝试提取")
    logger.info(
        "[BankReconciliationDetail][PDF] parse_status=success file=%s parsed_transaction_count=%s amount_status=%s fail_reason=",
        source_file,
        len(transactions),
        "完整" if all(tx.get("amount") is not None for tx in transactions) else "部分缺失",
    )
    return FileParseResult(
        source_file=source_file,
        sheet_name="PDF",
        bank_name=bank_name,
        header_row_no=header_line_no if header_detected else 0,
        header_col_start=1 if header_detected else 0,
        account=account,
        transactions=transactions,
        raw_summary=raw_summary,
        placeholder_cleaned_count=0,
        warnings=warnings,
    ), warnings


def _row_text(row: list[Any]) -> str:
    return " ".join(_text(cell) for cell in row if _text(cell))


def _detect_bank_format(rows: list[list[Any]], bank_hint: str, mapping: dict[str, int]) -> tuple[str, str]:
    if bank_hint == "上海银行":
        return "shanghai", "filename"
    if bank_hint == "工商银行":
        return "icbc", "filename"
    scanned_cells = [_text(cell) for row in rows[:30] for cell in row[:30] if _text(cell)]
    joined = " ".join(scanned_cells)
    header_fields = set(mapping)
    if "账户明细查询" in joined or "选择账号" in joined:
        return "shanghai", "sheet_marker"
    if "开户行" in joined and "上海银行" in joined:
        return "shanghai", "branch_marker"
    if {"transaction_id", "trade_time", "direction", "amount", "counterparty_name"} <= header_fields:
        return "shanghai", "header_fields"
    if {"voucher_no", "in_amount", "out_amount"} & header_fields:
        return "icbc", "header_fields"
    return ("shanghai", "amount_header") if "amount" in header_fields else ("generic", "unknown")


def _find_header(rows: list[list[Any]], bank_hint: str) -> tuple[int, dict[str, int], str, int]:
    aliases = {
        "transaction_id": ("交易流水号",),
        "trade_time": ("交易时间",),
        "accounting_date": ("记账日期",),
        "direction": ("交易方向", "借贷标志"),
        "amount": ("交易金额",),
        "balance": ("余额",),
        "counterparty_account": ("对手账号", "对方账号"),
        "counterparty_name": ("对手名称", "对方单位"),
        "counterparty_bank_no": ("对方行号",),
        "summary": ("摘要",),
        "purpose": ("交易用途", "用途"),
        "remark": ("备注", "附言"),
        "voucher_no": ("凭证号",),
        "in_amount": ("转入金额",),
        "out_amount": ("转出金额",),
    }
    best_idx = -1
    best_map: dict[str, int] = {}
    best_score = 0
    shanghai_required = {
        "交易流水号",
        "交易时间",
        "记账日期",
        "交易方向",
        "交易金额",
        "余额",
        "对手账号",
        "对手名称",
        "摘要",
        "交易用途",
        "备注",
    }
    for idx, row in enumerate(rows[:50]):
        mapping: dict[str, int] = {}
        shanghai_hits = 0
        for col, cell in enumerate(row[:50]):
            header = _compact(cell)
            if not header:
                continue
            if any(_compact(name) == header or _compact(name) in header for name in shanghai_required):
                shanghai_hits += 1
            for field, names in aliases.items():
                if field in mapping:
                    continue
                if any(_compact(name) == header or _compact(name) in header for name in names):
                    mapping[field] = col
        fields = set(mapping)
        score = len(fields) + (3 if "trade_time" in fields or "accounting_date" in fields else 0) + (3 if {"in_amount", "out_amount"} & fields or "amount" in fields else 0)
        if shanghai_hits >= 5:
            score += 20
        if score > best_score:
            best_idx, best_map, best_score = idx, mapping, score
    if best_score < 5:
        return -1, {}, "generic", 0
    bank_format, _ = _detect_bank_format(rows, bank_hint, best_map)
    header_col_start = min(best_map.values()) + 1 if best_map else 0
    return best_idx, best_map, bank_format, header_col_start


def _value_after_label(row: list[Any], label: str) -> str:
    compact_label = _compact(label).rstrip(":：")
    for idx, cell in enumerate(row):
        text = _text(cell)
        compact = _compact(text)
        if not compact:
            continue
        if compact.startswith(compact_label):
            for sep in (":", "："):
                if sep in text:
                    right = text.split(sep, 1)[1].strip()
                    if right:
                        return _text(right)
            if idx + 1 < len(row):
                return _text(row[idx + 1])
    return ""


def _find_label_index(row: list[Any], label: str) -> int:
    compact_label = _compact(label).rstrip(":：")
    for idx, cell in enumerate(row):
        compact = _compact(cell).rstrip(":：")
        if compact.startswith(compact_label):
            return idx
    return -1


def _next_value_after_index(row: list[Any], start_idx: int, stop_labels: tuple[str, ...] = ()) -> str:
    stop_compacts = tuple(_compact(label).rstrip(":：") for label in stop_labels)
    for idx in range(start_idx + 1, len(row)):
        value = _text(row[idx])
        compact = _compact(value).rstrip(":：")
        if not compact:
            continue
        if any(compact.startswith(label) for label in stop_compacts):
            return ""
        return value
    return ""


def _parse_meta(rows: list[list[Any]], sheet_name: str, source_file: str, bank_name: str) -> tuple[AccountInfo, dict[str, Any]]:
    account = AccountInfo(bank_name=bank_name or _bank_from_filename(source_file), source_file=source_file, sheet_name=sheet_name)
    summary: dict[str, Any] = {}
    for row in rows[:10]:
        row_joined = _row_text(row)
        if "记账日期" in row_joined:
            start, end = _date_range(row_joined)
            account.date_start = account.date_start or start
            account.date_end = account.date_end or end
        selected = _value_after_label(row, "选择账号")
        if selected:
            account.account_no = re.sub(r"\D", "", selected) or account.account_no
            selected_idx = _find_label_index(row, "选择账号")
            if selected_idx >= 0 and not account.account_name:
                possible_name = _next_value_after_index(row, selected_idx + 1, ("开户行", "币种"))
                if possible_name and not re.fullmatch(r"\d+", possible_name):
                    account.account_name = possible_name
        for label, field_name in (
            ("户名", "account_name"),
            ("开户行", "branch_name"),
            ("币种", "currency"),
        ):
            value = _value_after_label(row, label)
            if value:
                setattr(account, field_name, value)
        for label, key in (
            ("总笔数", "raw_transaction_count"),
            ("借方总笔数", "debit_count"),
            ("借方总金额", "debit_amount"),
            ("贷方总笔数", "credit_count"),
            ("贷方总金额", "credit_amount"),
        ):
            value = _value_after_label(row, label)
            if value:
                summary[key] = _money(value) if "金额" in label else _int(value)
    if not account.bank_name:
        account.bank_name = _bank_from_filename(source_file) or UNKNOWN
    if not account.currency:
        account.currency = "人民币"
    if not account.account_name and bank_name == "工商银行":
        account.account_confidence = "missing"
    return account, summary


def _row_get(row: list[Any], mapping: dict[str, int], key: str) -> Any:
    idx = mapping.get(key)
    return row[idx] if idx is not None and idx < len(row) else None


def _classify(tx: dict[str, Any], excluded_parties: dict[str, str]) -> None:
    joined = " ".join(_text(tx.get(key)) for key in ("counterparty_name", "summary", "purpose", "remark"))
    own_name = _normalize_party_name(tx.get("account_name"))
    counterparty = _normalize_party_name(tx.get("counterparty_name"))
    excluded_reason = excluded_parties.get(counterparty, "") if counterparty else ""
    if counterparty and own_name and counterparty == own_name:
        excluded_reason = "本方同名划转"
    tx["is_excluded_related_party"] = bool(excluded_reason)
    tx["excluded_reason"] = excluded_reason
    tx["related_party_type"] = excluded_reason.replace("往来", "") if excluded_reason and excluded_reason != "本方同名划转" else ("本方同名账户" if excluded_reason else "")
    tx["is_self_transfer"] = excluded_reason == "本方同名划转"
    tx["is_related_party_transfer"] = bool(excluded_reason)
    tx["is_noise"] = _is_noise_counterparty(tx.get("counterparty_name"), joined)
    tx["is_personal_counterparty"] = _is_personal_name(tx.get("counterparty_name"))
    tx["is_loan_related"] = _contains_any(joined, LOAN_KEYWORDS)
    tx["is_fee"] = _contains_any(joined, FEE_KEYWORDS)
    tx["is_salary"] = _contains_any(joined, SALARY_KEYWORDS)
    tx["is_tax"] = _contains_any(joined, TAX_KEYWORDS)
    tx["is_interest"] = _contains_any(joined, INTEREST_KEYWORDS)
    is_deposit = _contains_any(joined, DEPOSIT_KEYWORDS)
    has_operating_in = _contains_any(joined, OPERATING_IN_KEYWORDS)
    has_operating_out = _contains_any(joined, OPERATING_OUT_KEYWORDS)
    has_non_operating = _contains_any(joined, NON_OPERATING_KEYWORDS)
    is_org = _is_organization_counterparty(tx.get("counterparty_name"))
    if _is_unknown_counterparty(tx.get("counterparty_name")):
        tx["category"] = "未识别对手方"
        tx["is_noise"] = True
    elif tx["is_excluded_related_party"]:
        tx["category"] = "内部/关联方往来"
        tx["is_noise"] = False
    elif tx["is_personal_counterparty"]:
        tx["category"] = "个人往来"
    elif tx["is_loan_related"]:
        tx["category"] = "融资/贷款相关"
    elif tx["is_tax"]:
        tx["category"] = "税费/社保"
    elif tx["is_salary"]:
        tx["category"] = "人工成本/代发"
    elif tx["is_fee"] or tx["is_interest"]:
        tx["category"] = "手续费/利息"
    elif is_deposit:
        tx["category"] = "押金/保证金/退款"
    elif tx["is_noise"]:
        tx["category"] = "未识别对手方"
    elif tx.get("direction") == "in" and is_org and has_operating_in and not has_non_operating:
        tx["category"] = "经营性入账"
    elif tx.get("direction") == "out" and is_org and has_operating_out and not has_non_operating:
        tx["category"] = "经营性出账"
    else:
        tx["category"] = "非经营性往来"
    tx["is_operating_inflow"] = tx["category"] == "经营性入账"
    tx["is_operating_outflow"] = tx["category"] == "经营性出账"
    tx["confidence"] = "high" if tx["category"] in {"经营性入账", "经营性出账", "内部/关联方往来", "贷款相关"} else "medium"


def _parse_sheet(rows: list[list[Any]], sheet_name: str, source_file: str) -> FileParseResult | None:
    bank_hint = _bank_from_filename(source_file)
    max_cols = max((len(row) for row in rows), default=0)
    header_idx, mapping, bank_format, header_col_start = _find_header(rows, bank_hint)
    bank_format, bank_reason = _detect_bank_format(rows, bank_hint, mapping)
    field_map_for_log = {key: value + 1 for key, value in mapping.items()}
    logger.info(
        "[BankReconciliationDetail] scan file=%s sheet=%s range=%sx%s bank_hint=%s bank_format=%s bank_reason=%s header_row_no=%s header_col_start=%s header_col_start_letter=%s field_map=%s",
        source_file,
        sheet_name,
        len(rows),
        max_cols,
        bank_hint or UNKNOWN,
        bank_format,
        bank_reason,
        header_idx + 1 if header_idx >= 0 else 0,
        header_col_start,
        _column_letter(header_col_start),
        field_map_for_log,
    )
    if header_idx < 0 or not mapping:
        logger.warning(
            "[BankReconciliationDetail] parse_skip file=%s sheet=%s reason=header_not_found range=%sx%s",
            source_file,
            sheet_name,
            len(rows),
            max_cols,
        )
        return None
    bank_name = bank_hint or ("上海银行" if bank_format == "shanghai" else "工商银行" if bank_format == "icbc" else UNKNOWN)
    account, raw_summary = _parse_meta(rows, sheet_name, source_file, bank_name)
    placeholder_cleaned = 0
    transactions: list[dict[str, Any]] = []
    excluded_parties: dict[str, str] = {}
    _add_excluded_party(excluded_parties, account.account_name, "本方同名划转")
    empty_streak = 0
    for row_no, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(_text(cell) for cell in row):
            empty_streak += 1
            if empty_streak >= 8:
                break
            continue
        empty_streak = 0
        placeholder_cleaned += sum(1 for cell in row if _is_placeholder(cell))
        trade_time = _date_text(_row_get(row, mapping, "trade_time"), with_time=True)
        accounting_date = _date_text(_row_get(row, mapping, "accounting_date"))
        if not trade_time and accounting_date:
            trade_time = accounting_date
        direction_raw = _text(_row_get(row, mapping, "direction"))
        in_amount = _money(_row_get(row, mapping, "in_amount"))
        out_amount = _money(_row_get(row, mapping, "out_amount"))
        amount = _money(_row_get(row, mapping, "amount"))
        direction = ""
        if direction_raw in {"贷", "贷方"} or "入账" in direction_raw or "贷方" in direction_raw:
            direction = "in"
        elif direction_raw in {"借", "借方"} or "出账" in direction_raw or "借方" in direction_raw:
            direction = "out"
        elif in_amount is not None:
            direction = "in"
        elif out_amount is not None:
            direction = "out"
        final_amount = amount if amount is not None else (in_amount if direction == "in" else out_amount)
        if final_amount is None and not trade_time and not accounting_date:
            continue
        tx = {
            "transaction_id": _text(_row_get(row, mapping, "transaction_id")) or _text(_row_get(row, mapping, "voucher_no")),
            "source_file": source_file,
            "bank_name": bank_name,
            "account_name": account.account_name,
            "account_no": account.account_no,
            "trade_time": trade_time,
            "accounting_date": accounting_date or (trade_time[:10] if trade_time else ""),
            "direction": direction,
            "direction_name": "入账" if direction == "in" else "出账" if direction == "out" else UNKNOWN,
            "amount": final_amount,
            "balance": _money(_row_get(row, mapping, "balance")),
            "counterparty_name": _text(_row_get(row, mapping, "counterparty_name")),
            "counterparty_account": _text(_row_get(row, mapping, "counterparty_account")),
            "counterparty_bank_no": _text(_row_get(row, mapping, "counterparty_bank_no")),
            "summary": _text(_row_get(row, mapping, "summary")),
            "purpose": _text(_row_get(row, mapping, "purpose")),
            "remark": _text(_row_get(row, mapping, "remark")),
            "voucher_no": _text(_row_get(row, mapping, "voucher_no")),
            "raw_row_no": row_no,
            "warning": "" if final_amount is not None else "金额未识别",
        }
        _classify(tx, excluded_parties)
        transactions.append(tx)
    dates = sorted(tx["accounting_date"] for tx in transactions if tx.get("accounting_date"))
    if dates:
        account.date_start = account.date_start or dates[0]
        account.date_end = account.date_end or dates[-1]
    if not raw_summary.get("raw_transaction_count"):
        raw_summary["raw_transaction_count"] = len(transactions)
    logger.info(
        "[BankReconciliationDetail] parsed file=%s sheet=%s bank=%s header_row_no=%s header_col_start=%s header_col_start_letter=%s read_rows=%s parsed_transaction_count=%s amount_status=%s",
        source_file,
        sheet_name,
        bank_name,
        header_idx + 1,
        header_col_start,
        _column_letter(header_col_start),
        max(0, len(rows) - header_idx - 1),
        len(transactions),
        "完整" if transactions and all(tx.get("amount") is not None for tx in transactions) else "部分缺失" if transactions else "未识别",
    )
    return FileParseResult(
        source_file=source_file,
        sheet_name=sheet_name,
        bank_name=bank_name,
        header_row_no=header_idx + 1,
        header_col_start=header_col_start,
        account=account,
        transactions=transactions,
        raw_summary=raw_summary,
        placeholder_cleaned_count=placeholder_cleaned,
        warnings=["已清理占位值 17"] if placeholder_cleaned else [],
    )


def _dedupe(transactions: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    seen: set[str] = set()
    unique: list[dict[str, Any]] = []
    for tx in transactions:
        signature = "|".join(
            _text(tx.get(key))
            for key in (
                "bank_name",
                "account_no",
                "trade_time",
                "accounting_date",
                "direction",
                "counterparty_account",
                "counterparty_name",
                "summary",
                "purpose",
                "remark",
            )
        )
        signature += f"|{_fmt_money(tx.get('amount'))}|{_fmt_money(tx.get('balance'))}"
        digest = hashlib.sha1(signature.encode("utf-8")).hexdigest()
        if digest in seen:
            continue
        seen.add(digest)
        unique.append(tx)
    return unique, len(transactions) - len(unique)


def _sum(transactions: list[dict[str, Any]], predicate: Any) -> Decimal:
    total = Decimal("0")
    for tx in transactions:
        if predicate(tx):
            amount = _money(tx.get("amount"))
            if amount is not None:
                total += amount
    return total


def _period(transactions: list[dict[str, Any]], accounts: list[dict[str, Any]]) -> tuple[str, str]:
    dates = [tx.get("accounting_date") for tx in transactions if tx.get("accounting_date")]
    dates.extend(account.get("date_start") for account in accounts if account.get("date_start"))
    dates.extend(account.get("date_end") for account in accounts if account.get("date_end"))
    dates = sorted(str(item) for item in dates if item)
    return (dates[0], dates[-1]) if dates else ("", "")


def _account_key(account: dict[str, Any]) -> str:
    return "|".join([str(account.get("bank_name") or ""), str(account.get("account_no") or ""), str(account.get("source_file") or ""), str(account.get("sheet_name") or "")])


def _finalize_top_counter(counter: dict[str, dict[str, Any]]) -> list[tuple[str, dict[str, Any]]]:
    finalized: list[tuple[str, dict[str, Any]]] = []
    for name, item in counter.items():
        evidence_map = item.get("evidence") or {}
        evidence_items = sorted(evidence_map.items(), key=lambda pair: pair[1], reverse=True)
        evidence = "、".join(str(key)[:20] for key, _ in evidence_items[:3] if key)
        if not evidence:
            continue
        finalized.append(
            (
                name,
                {
                    "amount": item.get("amount") or Decimal("0"),
                    "count": item.get("count") or 0,
                    "operating": item.get("operating") or 0,
                    "evidence": evidence,
                },
            )
        )
    return sorted(finalized, key=lambda item: item[1]["amount"], reverse=True)[:10]


def _aggregate(file_results: list[FileParseResult], excluded_parties: dict[str, str] | None = None) -> dict[str, Any]:
    all_excluded_parties = dict(excluded_parties or {})
    own_display_names: list[str] = []
    for result in file_results:
        normalized_name = _normalize_party_name(result.account.account_name)
        if normalized_name:
            all_excluded_parties.setdefault(normalized_name, "本方同名划转")
            if result.account.account_name not in own_display_names:
                own_display_names.append(result.account.account_name)
    for name, reason in all_excluded_parties.items():
        if reason == "本方同名划转" and name and name not in own_display_names:
            own_display_names.append(name)
    inherited_account_name = own_display_names[0] if own_display_names else ""
    raw_transactions = [tx for result in file_results for tx in result.transactions]
    for tx in raw_transactions:
        _classify(tx, all_excluded_parties)
    transactions, duplicate_count = _dedupe(raw_transactions)
    account_map: dict[str, dict[str, Any]] = {}
    for result in file_results:
        account = result.account.__dict__.copy()
        if not account.get("account_name") and inherited_account_name:
            account["account_name"] = inherited_account_name
            account["account_confidence"] = "inherited"
        key = _account_key(account)
        account_map.setdefault(key, {**account, "raw_count": 0, "deduped_count": 0})
        account_map[key]["raw_count"] += len(result.transactions)
    for tx in transactions:
        key = "|".join([str(tx.get("bank_name") or ""), str(tx.get("account_no") or ""), str(tx.get("source_file") or ""), str(next((r.sheet_name for r in file_results if r.source_file == tx.get("source_file")), ""))])
        if key in account_map:
            account_map[key]["deduped_count"] += 1
    accounts = list(account_map.values())
    start, end = _period(transactions, accounts)
    monthly: dict[str, dict[str, Any]] = defaultdict(lambda: {"in": Decimal("0"), "out": Decimal("0"), "op_in": Decimal("0"), "op_out": Decimal("0"), "count": 0, "op_count": 0})
    in_counter: dict[str, dict[str, Any]] = defaultdict(lambda: {"amount": Decimal("0"), "count": 0, "operating": 0, "evidence": defaultdict(lambda: Decimal("0"))})
    out_counter: dict[str, dict[str, Any]] = defaultdict(lambda: {"amount": Decimal("0"), "count": 0, "operating": 0, "evidence": defaultdict(lambda: Decimal("0"))})
    for tx in transactions:
        month = str(tx.get("accounting_date") or "")[:7] or UNKNOWN
        amount = _money(tx.get("amount")) or Decimal("0")
        monthly[month]["count"] += 1
        if tx.get("direction") == "in":
            monthly[month]["in"] += amount
            if tx.get("is_operating_inflow"):
                monthly[month]["op_in"] += amount
                monthly[month]["op_count"] += 1
            if _is_top_eligible_operating_tx(tx, "in"):
                evidence = _operating_evidence(tx, "in")
                if not evidence:
                    continue
                item = in_counter[_display(tx.get("counterparty_name"))]
                item["amount"] += amount
                item["count"] += 1
                item["operating"] += 1
                item["evidence"][evidence] += amount
        elif tx.get("direction") == "out":
            monthly[month]["out"] += amount
            if tx.get("is_operating_outflow"):
                monthly[month]["op_out"] += amount
                monthly[month]["op_count"] += 1
            if _is_top_eligible_operating_tx(tx, "out"):
                evidence = _operating_evidence(tx, "out")
                if not evidence:
                    continue
                item = out_counter[_display(tx.get("counterparty_name"))]
                item["amount"] += amount
                item["count"] += 1
                item["operating"] += 1
                item["evidence"][evidence] += amount
    return {
        "doc_type": DOC_TYPE,
        "doc_type_name": DOC_TYPE_NAME,
        "document_type": DOC_TYPE,
        "document_type_code": DOC_TYPE,
        "document_type_name": DOC_TYPE_NAME,
        "agent_type": AGENT_TYPE,
        "skill_name": SKILL_NAME,
        "schema_version": SCHEMA_VERSION,
        "extraction_status": "success" if transactions else "failed",
        "files": [
            {
                "source_file": r.source_file,
                "sheet_name": r.sheet_name,
                "bank_name": r.bank_name,
                "header_row_no": r.header_row_no,
                "header_col_start": r.header_col_start,
                "transaction_count": len(r.transactions),
                "amount_status": "完整" if all(tx.get("amount") is not None for tx in r.transactions) else "部分缺失",
                "date_start": r.account.date_start,
                "date_end": r.account.date_end,
                "status": r.status,
                "warnings": r.warnings,
                "placeholder_cleaned_count": r.placeholder_cleaned_count,
                "raw_summary": r.raw_summary,
            }
            for r in file_results
        ],
        "accounts": accounts,
        "summary": {
            "file_count": len({r.source_file for r in file_results}),
            "account_count": len(accounts),
            "date_start": start,
            "date_end": end,
            "raw_transaction_count": len(raw_transactions),
            "deduped_transaction_count": len(transactions),
            "duplicate_transaction_count": duplicate_count,
            "in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in"),
            "out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out"),
            "self_transfer_in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in" and tx.get("is_self_transfer")),
            "self_transfer_out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out" and tx.get("is_self_transfer")),
            "in_amount_excluding_self_transfer": _sum(transactions, lambda tx: tx.get("direction") == "in" and not tx.get("is_self_transfer")),
            "out_amount_excluding_self_transfer": _sum(transactions, lambda tx: tx.get("direction") == "out" and not tx.get("is_self_transfer")),
            "excluded_related_transaction_count": sum(1 for tx in transactions if tx.get("is_excluded_related_party") or tx.get("is_related_party_transfer")),
            "excluded_related_in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in" and (tx.get("is_excluded_related_party") or tx.get("is_related_party_transfer"))),
            "excluded_related_out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out" and (tx.get("is_excluded_related_party") or tx.get("is_related_party_transfer"))),
            "in_amount_excluding_excluded_related": _sum(transactions, lambda tx: tx.get("direction") == "in" and not (tx.get("is_excluded_related_party") or tx.get("is_related_party_transfer"))),
            "out_amount_excluding_excluded_related": _sum(transactions, lambda tx: tx.get("direction") == "out" and not (tx.get("is_excluded_related_party") or tx.get("is_related_party_transfer"))),
            "personal_transaction_count": sum(1 for tx in transactions if tx.get("category") == "个人往来"),
            "personal_in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in" and tx.get("category") == "个人往来"),
            "personal_out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out" and tx.get("category") == "个人往来"),
            "loan_interest_transaction_count": sum(1 for tx in transactions if tx.get("category") == "融资/贷款相关" or (tx.get("category") == "手续费/利息" and tx.get("is_interest") and not tx.get("is_fee"))),
            "loan_interest_in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in" and (tx.get("category") == "融资/贷款相关" or (tx.get("category") == "手续费/利息" and tx.get("is_interest") and not tx.get("is_fee")))),
            "loan_interest_out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out" and (tx.get("category") == "融资/贷款相关" or (tx.get("category") == "手续费/利息" and tx.get("is_interest") and not tx.get("is_fee")))),
            "salary_tax_fee_transaction_count": sum(1 for tx in transactions if tx.get("category") in {"税费/社保", "人工成本/代发"} or (tx.get("category") == "手续费/利息" and tx.get("is_fee"))),
            "salary_tax_fee_in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in" and (tx.get("category") in {"税费/社保", "人工成本/代发"} or (tx.get("category") == "手续费/利息" and tx.get("is_fee")))),
            "salary_tax_fee_out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out" and (tx.get("category") in {"税费/社保", "人工成本/代发"} or (tx.get("category") == "手续费/利息" and tx.get("is_fee")))),
            "noise_transaction_count": sum(1 for tx in transactions if tx.get("category") == "未识别对手方"),
            "noise_in_amount": _sum(transactions, lambda tx: tx.get("direction") == "in" and tx.get("category") == "未识别对手方"),
            "noise_out_amount": _sum(transactions, lambda tx: tx.get("direction") == "out" and tx.get("category") == "未识别对手方"),
            "operating_in_amount": _sum(transactions, lambda tx: tx.get("is_operating_inflow")),
            "operating_out_amount": _sum(transactions, lambda tx: tx.get("is_operating_outflow")),
            "amount_completeness": "完整" if transactions and all(tx.get("amount") is not None for tx in transactions) else "部分缺失" if transactions else "未识别",
            "aggregation_status": "可用" if transactions else "不可用",
        },
        "monthly": dict(sorted(monthly.items())),
        "top_in": _finalize_top_counter(in_counter),
        "top_out": _finalize_top_counter(out_counter),
        "transactions": transactions,
        "warnings": list(dict.fromkeys(w for r in file_results for w in r.warnings)),
    }


def _render_markdown(data: dict[str, Any]) -> str:
    summary = data.get("summary") or {}
    transactions = data.get("transactions") or []
    lines: list[str] = [
        "## 银行对账明细",
        "",
        f"- 资料类型：{DOC_TYPE_NAME}",
        f"- 提取状态：{'成功' if summary.get('deduped_transaction_count') else '失败'}",
        f"- 覆盖文件数：{_fmt_count(summary.get('file_count'))} 份",
        f"- 已识别银行账户数：{_fmt_count(summary.get('account_count'))} 个",
        f"- 覆盖时间范围：{_display(summary.get('date_start'))} 至 {_display(summary.get('date_end'))}",
        f"- 原始交易笔数：{_fmt_count(summary.get('raw_transaction_count'))}",
        f"- 去重后交易笔数：{_fmt_count(summary.get('deduped_transaction_count'))}",
        f"- 重复交易笔数：{_fmt_count(summary.get('duplicate_transaction_count'))}",
        f"- 金额识别完整度：{_display(summary.get('amount_completeness'), '未识别')}",
        f"- 聚合状态：{_display(summary.get('aggregation_status'), '不可用')}",
        "",
        "### 文件解析质量清单",
        "",
        "| 序号 | 来源文件 | 银行 | 工作表 | 表头识别 | 明细笔数 | 金额识别 | 时间范围 | 状态 | 提示 |",
        "|---|---|---|---|---|---:|---|---|---|---|",
    ]
    for idx, item in enumerate(data.get("files") or [], start=1):
        tip = "；".join(item.get("warnings") or []) or "-"
        lines.append(f"| {idx} | {_display(item.get('source_file'))} | {_display(item.get('bank_name'))} | {_display(item.get('sheet_name'))} | {'已识别' if item.get('header_row_no') else '未识别'} | {_fmt_count(item.get('transaction_count'))} | {_display(item.get('amount_status'), '未识别')} | {_display(item.get('date_start'))} 至 {_display(item.get('date_end'))} | {_display(item.get('status'), '失败')} | {_display(tip, '-')} |")
    lines += ["", "### 银行账户汇总", "", "| 序号 | 银行 | 户名 | 账号 | 开户行 | 币种 | 时间范围 | 交易笔数 |", "|---|---|---|---|---|---|---|---:|"]
    for idx, account in enumerate(data.get("accounts") or [], start=1):
        lines.append(f"| {idx} | {_display(account.get('bank_name'))} | {_display(account.get('account_name'))} | {_display(account.get('account_no'))} | {_display(account.get('branch_name'))} | {_display(account.get('currency'), '人民币')} | {_display(account.get('date_start'))} 至 {_display(account.get('date_end'))} | {_fmt_count(account.get('deduped_count') or account.get('raw_count'))} |")
    net = (_money(summary.get("in_amount")) or Decimal("0")) - (_money(summary.get("out_amount")) or Decimal("0"))
    op_net = (_money(summary.get("operating_in_amount")) or Decimal("0")) - (_money(summary.get("operating_out_amount")) or Decimal("0"))
    lines += [
        "",
        "### 资金总览",
        "",
        "| 项目 | 金额/数量 |",
        "|---|---:|",
        f"| 原始入账金额 | {_fmt_money(summary.get('in_amount'))} |",
        f"| 原始出账金额 | {_fmt_money(summary.get('out_amount'))} |",
        f"| 资金净流入 | {_fmt_money(net)} |",
        f"| 本方同名划转入账 | {_fmt_money(summary.get('self_transfer_in_amount'))} |",
        f"| 本方同名划转出账 | {_fmt_money(summary.get('self_transfer_out_amount'))} |",
        f"| 有效经营入账 | {_fmt_money(summary.get('operating_in_amount'))} |",
        f"| 有效经营出账 | {_fmt_money(summary.get('operating_out_amount'))} |",
        f"| 经营净流入 | {_fmt_money(op_net)} |",
        "",
        "### 月度汇总",
        "",
        "| 月份 | 入账金额 | 出账金额 | 净流入 | 经营性入账 | 经营性出账 | 经营净流入 | 交易笔数 |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for month, item in (data.get("monthly") or {}).items():
        net_month = item["in"] - item["out"]
        op_net_month = item["op_in"] - item["op_out"]
        lines.append(f"| {month} | {_fmt_money(item['in'])} | {_fmt_money(item['out'])} | {_fmt_money(net_month)} | {_fmt_money(item['op_in'])} | {_fmt_money(item['op_out'])} | {_fmt_money(op_net_month)} | {_fmt_count(item['count'])} |")
    for title, key, direction in (("主要入账对象 TOP10", "top_in", "入账"), ("主要出账对象 TOP10", "top_out", "出账")):
        lines += ["", f"### {title}", "", f"| 排名 | 对方户名 | {direction}金额 | {direction}笔数 | 是否经营性 | 备注 |", "|---|---|---:|---:|---|---|"]
        for idx, (name, item) in enumerate(data.get(key) or [], start=1):
            operating = "是" if item.get("operating") else "否"
            lines.append(f"| {idx} | {_display(name)} | {_fmt_money(item.get('amount'))} | {_fmt_count(item.get('count'))} | {operating} | - |")
        if not data.get(key):
            lines.append("| - | 无 | 0.00 | 0 | 否 | 无 |")
    self_transfers = [tx for tx in transactions if tx.get("is_self_transfer")]
    lines += ["", "### 本方同名划转识别", ""]
    if self_transfers:
        lines += ["| 序号 | 日期 | 银行 | 对方户名 | 方向 | 金额 | 摘要/用途 | 处理结果 |", "|---|---|---|---|---|---:|---|---|"]
        for idx, tx in enumerate(self_transfers[:20], start=1):
            result = "已剔除经营收入" if tx.get("direction") == "in" else "已剔除经营支出"
            lines.append(f"| {idx} | {_display(tx.get('accounting_date'))} | {_display(tx.get('bank_name'))} | {_display(tx.get('counterparty_name'))} | {_display(tx.get('direction_name'))} | {_fmt_money(tx.get('amount'))} | {_display(tx.get('summary') or tx.get('purpose'), '无')} | {result} |")
    else:
        lines.append("- 本方同名划转：未发现")
    placeholder_count = sum(int(item.get("placeholder_cleaned_count") or 0) for item in data.get("files") or [])
    amount_missing = any(tx.get("amount") is None for tx in transactions)
    date_missing = any(not tx.get("accounting_date") for tx in transactions)
    large_in = any((_money(tx.get("amount")) or Decimal("0")) >= Decimal("1000000") and tx.get("direction") == "in" for tx in transactions)
    lines += [
        "",
        "### 异常和提示",
        "",
        f"- 是否存在金额缺失：{'是' if amount_missing else '否'}",
        f"- 是否存在日期缺失：{'是' if date_missing else '否'}",
        f"- 是否存在占位值清理：{'是' if placeholder_count else '否'}",
        f"- 是否存在大额集中收款：{'是' if large_in else '否'}",
        f"- 是否存在本方同名大额往来：{'是' if self_transfers else '否'}",
        f"- 是否存在贷款/利息/手续费类交易：{'是' if any(tx.get('is_loan_related') or tx.get('is_interest') or tx.get('is_fee') for tx in transactions) else '否'}",
        "",
        "### 交易明细样例",
        "",
        "| 日期 | 银行 | 方向 | 金额 | 对方户名 | 摘要 | 用途 | 分类 |",
        "|---|---|---|---:|---|---|---|---|",
    ]
    for tx in transactions[:50]:
        lines.append(f"| {_display(tx.get('accounting_date'))} | {_display(tx.get('bank_name'))} | {_display(tx.get('direction_name'))} | {_fmt_money(tx.get('amount'))} | {_display(tx.get('counterparty_name'))} | {_display(tx.get('summary'), '无')} | {_display(tx.get('purpose'), '无')} | {_display(tx.get('category'), '未分类')} |")
    if not transactions:
        lines.append("| 无 | 无 | 无 | 0.00 | 无 | 无 | 无 | 无 |")
    markdown = "\n".join(lines)
    for forbidden in ("raw_result", "normalized_data", "fields:", "transactions:", "null", "undefined", "None", "{}", "[]", "| 17 |", "> 17 <"):
        markdown = markdown.replace(forbidden, "")
    return markdown


def _render_compact_display_markdown(data: dict[str, Any]) -> str:
    summary = data.get("summary") or {}
    files = data.get("files") or []
    accounts = data.get("accounts") or []
    primary_account = accounts[0] if accounts else {}
    transactions = data.get("transactions") or []
    file_count = summary.get("file_count") or len(files)
    source_file = _display(files[0].get("source_file")) if len(files) == 1 else f"{_fmt_count(file_count)} 份文件"
    banks = sorted({_display(item.get("bank_name"), "") for item in files if _display(item.get("bank_name"), "")})
    bank_name = banks[0] if len(banks) == 1 else ("多家银行" if banks else UNKNOWN)
    status = "成功" if summary.get("deduped_transaction_count") else "失败"
    in_amount = _money(summary.get("in_amount")) or Decimal("0")
    out_amount = _money(summary.get("out_amount")) or Decimal("0")
    excluded_related_in = _money(summary.get("excluded_related_in_amount")) or _money(summary.get("self_transfer_in_amount")) or Decimal("0")
    excluded_related_out = _money(summary.get("excluded_related_out_amount")) or _money(summary.get("self_transfer_out_amount")) or Decimal("0")
    personal_in = _money(summary.get("personal_in_amount")) or Decimal("0")
    personal_out = _money(summary.get("personal_out_amount")) or Decimal("0")
    loan_interest_in = _money(summary.get("loan_interest_in_amount")) or Decimal("0")
    loan_interest_out = _money(summary.get("loan_interest_out_amount")) or Decimal("0")
    salary_tax_fee_in = _money(summary.get("salary_tax_fee_in_amount")) or Decimal("0")
    salary_tax_fee_out = _money(summary.get("salary_tax_fee_out_amount")) or Decimal("0")
    noise_in = _money(summary.get("noise_in_amount")) or Decimal("0")
    noise_out = _money(summary.get("noise_out_amount")) or Decimal("0")
    in_excluding_related = in_amount - excluded_related_in - personal_in - loan_interest_in - noise_in
    out_excluding_related = out_amount - excluded_related_out - personal_out - loan_interest_out - salary_tax_fee_out - noise_out
    excluded_related_count = int(summary.get("excluded_related_transaction_count") or 0)
    personal_count = int(summary.get("personal_transaction_count") or 0)
    loan_interest_count = int(summary.get("loan_interest_transaction_count") or 0)
    salary_tax_fee_count = int(summary.get("salary_tax_fee_transaction_count") or 0)
    noise_count = int(summary.get("noise_transaction_count") or 0)
    operating_in = _money(summary.get("operating_in_amount")) or Decimal("0")
    operating_out = _money(summary.get("operating_out_amount")) or Decimal("0")
    net = in_amount - out_amount
    op_net = operating_in - operating_out
    operating_in_ratio = (operating_in / in_amount) if in_amount else Decimal("0")
    operating_out_ratio = (operating_out / out_amount) if out_amount else Decimal("0")
    non_operating_in = in_amount - operating_in
    non_operating_out = out_amount - operating_out
    categorized_excluded_in = excluded_related_in + personal_in + loan_interest_in + salary_tax_fee_in + noise_in
    categorized_excluded_out = excluded_related_out + personal_out + loan_interest_out + salary_tax_fee_out + noise_out
    other_non_operating_in = non_operating_in - categorized_excluded_in
    other_non_operating_out = non_operating_out - categorized_excluded_out
    if abs(other_non_operating_in) < Decimal("0.01"):
        other_non_operating_in = Decimal("0")
    if abs(other_non_operating_out) < Decimal("0.01"):
        other_non_operating_out = Decimal("0")
    category_in_sum = categorized_excluded_in + other_non_operating_in
    category_out_sum = categorized_excluded_out + other_non_operating_out
    logger.info(
        "bank_reconciliation excluded reconciliation: excluded_in_total=%s category_in_sum=%s diff_in=%s excluded_out_total=%s category_out_sum=%s diff_out=%s",
        non_operating_in,
        category_in_sum,
        non_operating_in - category_in_sum,
        non_operating_out,
        category_out_sum,
        non_operating_out - category_out_sum,
    )
    base_amount = max(in_amount, out_amount, Decimal("1"))
    excluded_related_rows = [tx for tx in transactions if tx.get("is_excluded_related_party") or tx.get("is_self_transfer") or tx.get("is_related_party_transfer")]
    loan_like = [tx for tx in transactions if tx.get("is_loan_related") or tx.get("is_interest")]
    fee_like = [tx for tx in transactions if tx.get("is_fee")]
    personal_like = [tx for tx in transactions if tx.get("is_personal_counterparty") and not (tx.get("is_excluded_related_party") or tx.get("is_related_party_transfer"))]
    noise_like = [tx for tx in transactions if tx.get("is_noise")]
    operating_out_rows = [tx for tx in transactions if tx.get("is_operating_outflow")]

    lines: list[str] = [
        "## 银行对账明细",
        "",
        f"- 资料类型：{DOC_TYPE_NAME}",
        f"- 来源文件：{source_file}",
        f"- 提取状态：{status}",
        f"- 银行名称：{bank_name}",
        f"- 户名：{_display(primary_account.get('account_name'))}",
        f"- 账号：{_display(primary_account.get('account_no'))}",
        f"- 开户行：{_display(primary_account.get('branch_name'))}",
        f"- 覆盖时间：{_display(summary.get('date_start'))} 至 {_display(summary.get('date_end'))}",
        f"- 交易笔数：{_fmt_count(summary.get('deduped_transaction_count'))} 笔",
        f"- 金额识别：{_display(summary.get('amount_completeness'), UNKNOWN)}",
        "",
        "### 核心资金概览",
        "",
        "| 项目 | 金额/数量 |",
        "|---|---:|",
        f"| 原始入账总额 | {_fmt_money(in_amount)} |",
        f"| 原始出账总额 | {_fmt_money(out_amount)} |",
        f"| 原始净流入 | {_fmt_money(net)} |",
        f"| 剔除非经营入账 | {_fmt_money(non_operating_in)} |",
        f"| 剔除非经营出账 | {_fmt_money(non_operating_out)} |",
        f"| 有效经营入账 | {_fmt_money(operating_in)} |",
        f"| 有效经营出账 | {_fmt_money(operating_out)} |",
        f"| 经营净流入 | {_fmt_money(op_net)} |",
        f"| 有效经营入账占原始入账比例 | {_fmt_percent(operating_in, in_amount)} |",
        f"| 有效经营出账占原始出账比例 | {_fmt_percent(operating_out, out_amount)} |",
        "",
        "### 经营判断",
        "",
    ]
    if abs(net) / base_amount < Decimal("0.03"):
        lines.append("- 原始入账与出账基本持平，账面资金沉淀较少。")
    elif net > 0:
        lines.append("- 整体资金呈净流入，需要结合交易对手和用途判断回款质量。")
    else:
        lines.append("- 整体资金呈净流出，需关注持续支出对现金流的压力。")
    if op_net < 0:
        lines.append(f"- 剔除内部/关联方、贷款融资、手续费、未识别及其他非经营交易后，经营净流入为 {_fmt_money(op_net)}，经营性现金流为负。")
    elif op_net > 0:
        lines.append(f"- 剔除非经营交易后，经营净流入为 {_fmt_money(op_net)}，说明经营回款对经营支出有一定覆盖。")
    else:
        lines.append("- 有效经营入账与经营出账基本持平。")
    ratio_level = "偏低" if operating_in_ratio < Decimal("0.5") else "相对较高"
    lines.append(f"- 有效经营入账占原始入账比例为 {_fmt_percent(operating_in, in_amount)}，可采信经营回款占比{ratio_level}。")
    if operating_in < operating_out:
        lines.append("- 有效经营入账低于有效经营出账，需要关注真实回款能力和经营支出压力。")
    elif operating_in > operating_out:
        lines.append("- 有效经营入账高于有效经营出账，经营性现金流表现相对较好。")
    lines.append("- 当前主要经营客户/供应商仅统计有明确经营摘要或用途的交易。")

    if excluded_related_rows or personal_like or loan_like or fee_like or noise_like:
        other_non_operating_count = sum(
            1
            for tx in transactions
            if not tx.get("is_operating_inflow")
            and not tx.get("is_operating_outflow")
            and tx.get("category") in {"非经营性往来", "押金/保证金/退款"}
        )
        lines += [
            "",
            "### 非经营性及噪音剔除说明",
            "",
            "| 剔除类型 | 入账金额 | 出账金额 | 笔数 | 说明 |",
            "|---|---:|---:|---:|---|",
            f"| 内部/关联方往来 | {_fmt_money(excluded_related_in)} | {_fmt_money(excluded_related_out)} | {_fmt_count(excluded_related_count or len(excluded_related_rows))} | 本方同名、法人、股东、关联方 |",
            f"| 个人往来 | {_fmt_money(personal_in)} | {_fmt_money(personal_out)} | {_fmt_count(personal_count or len(personal_like))} | 非企业经营对手方 |",
            f"| 贷款/融资/利息类 | {_fmt_money(loan_interest_in)} | {_fmt_money(loan_interest_out)} | {_fmt_count(loan_interest_count or len(loan_like))} | 放款、还款、融资租赁、利息 |",
            f"| 工资/代发/税费/手续费类 | {_fmt_money(salary_tax_fee_in)} | {_fmt_money(salary_tax_fee_out)} | {_fmt_count(salary_tax_fee_count)} | 代发、税费、手续费等 |",
            f"| 未识别及噪音账户 | {_fmt_money(noise_in)} | {_fmt_money(noise_out)} | {_fmt_count(noise_count or len(noise_like))} | 空户名、银行系统账户等 |",
            f"| 其他非经营往来 | {_fmt_money(other_non_operating_in)} | {_fmt_money(other_non_operating_out)} | {_fmt_count(other_non_operating_count)} | 往来款、借还款、保证金、弱证据交易 |",
            "",
            "- 上述剔除分类合计应与核心资金概览中的“剔除非经营入账/出账”一致。",
        ]

    lines += [
        "",
        "### 月度经营资金变化",
        "",
        "| 月份 | 有效经营入账 | 有效经营出账 | 经营净流入 | 经营交易笔数 |",
        "|---|---:|---:|---:|---:|",
    ]
    rendered_operating_months = 0
    for month, item in (data.get("monthly") or {}).items():
        if not item.get("op_in") and not item.get("op_out") and not item.get("op_count"):
            continue
        net_month = item["op_in"] - item["op_out"]
        lines.append(f"| {month} | {_fmt_money(item['op_in'])} | {_fmt_money(item['op_out'])} | {_fmt_money(net_month)} | {_fmt_count(item.get('op_count'))} |")
        rendered_operating_months += 1
    if not rendered_operating_months:
        lines.append("| 未识别 | 0.00 | 0.00 | 0.00 | 0 |")

    def judgment(name: str, item: dict[str, Any], direction: str) -> str:
        return "经营性入账" if direction == "in" else "经营性出账"

    lines += [
        "",
        "### 主要经营入账来源",
        "",
        "| 排名 | 对方户名 | 入账金额 | 笔数 | 判断 | 经营依据 |",
        "|---|---|---:|---:|---|---|",
    ]
    for idx, (name, item) in enumerate(data.get("top_in") or [], start=1):
        lines.append(f"| {idx} | {_display(name)} | {_fmt_money(item.get('amount'))} | {_fmt_count(item.get('count'))} | {judgment(name, item, 'in')} | {_display(item.get('evidence'), '无')} |")
    if not data.get("top_in"):
        lines.append("| - | 无 | 0.00 | 0 | 无 | 无 |")

    lines += [
        "",
        "### 主要经营出账对象",
        "",
        "| 排名 | 对方户名 | 出账金额 | 笔数 | 判断 | 经营依据 |",
        "|---|---|---:|---:|---|---|",
    ]
    for idx, (name, item) in enumerate(data.get("top_out") or [], start=1):
        lines.append(f"| {idx} | {_display(name)} | {_fmt_money(item.get('amount'))} | {_fmt_count(item.get('count'))} | {judgment(name, item, 'out')} | {_display(item.get('evidence'), '无')} |")
    if not data.get("top_out"):
        lines.append("| - | 无 | 0.00 | 0 | 无 | 无 |")

    lines += ["", "### 风险提示", ""]
    risks: list[str] = []
    if op_net < 0:
        risks.append("剔除非经营交易后经营净流入为负，融资分析中不宜直接按原始流水总额判断还款能力。")
    if operating_in_ratio < Decimal("0.5") and in_amount:
        risks.append("有效经营入账占原始入账比例低于 50%，可采信经营回款占比偏低。")
    if non_operating_in > 0:
        risks.append("原始入账中存在较多内部往来、融资或非经营资金，需要结合发票、合同、应收账款核验真实销售回款。")
    if excluded_related_rows:
        risks.append("已识别内部/关联方往来，融资分析中应单独剔除，不作为经营回款或经营采购支出。")
    if not risks:
        risks.append("未识别到明显集中风险，仍建议结合合同、发票和回款周期复核。")
    lines.extend(f"- {risk}" for risk in list(dict.fromkeys(risks)))
    markdown = "\n".join(lines)
    for forbidden in ("raw_result", "normalized_data", "fields:", "transactions:", "null", "undefined", "None", "{}", "[]", "| 17 |", "> 17 <"):
        markdown = markdown.replace(forbidden, "")
    return markdown


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, tuple):
        return [_json_safe(item) for item in value]
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    return value


def _failure_result(reason: str, suggestion: str = "请检查上传文件是否成功保存，或检查调度层是否将文件路径传入银行对账明细 Agent。") -> dict[str, Any]:
    markdown = "\n".join(
        [
            "## 银行对账明细",
            "",
            f"- 资料类型：{DOC_TYPE_NAME}",
            "- 提取状态：失败",
            f"- 失败原因：{reason}",
            f"- 处理建议：{suggestion}",
        ]
    )
    return {
        "doc_type": DOC_TYPE,
        "doc_type_name": DOC_TYPE_NAME,
        "document_type": DOC_TYPE,
        "document_type_code": DOC_TYPE,
        "document_type_name": DOC_TYPE_NAME,
        "agent_type": AGENT_TYPE,
        "skill_name": SKILL_NAME,
        "schema_version": SCHEMA_VERSION,
        "extraction_status": "failed",
        "failure_reason": reason,
        "warnings": [reason],
        "files": [],
        "accounts": [],
        "summary": {
            "file_count": 0,
            "account_count": 0,
            "raw_transaction_count": 0,
            "deduped_transaction_count": 0,
            "duplicate_transaction_count": 0,
            "amount_completeness": "未识别",
            "aggregation_status": "不可用",
        },
        "monthly": {},
        "top_in": [],
        "top_out": [],
        "transactions": [],
        "display_markdown": markdown,
        "markdown": markdown,
        "markdown_summary": markdown,
        "report_markdown": markdown,
    }


def _normalize_input_files(files: Any) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    if not isinstance(files, list):
        return normalized
    for item in files:
        if not isinstance(item, dict):
            continue
        file_path = str(item.get("file_path") or item.get("path") or item.get("filePath") or "").strip()
        file_name = str(item.get("file_name") or item.get("filename") or item.get("fileName") or Path(file_path).name).strip()
        if file_path or file_name:
            normalized_item: dict[str, Any] = {"file_path": file_path, "file_name": file_name}
            for key in ("raw_pages", "pages", "text", "content", "mime_type"):
                if key in item:
                    normalized_item[key] = item.get(key)
            normalized.append(normalized_item)
    return normalized


def parse_bank_reconciliation_files(files: list[dict[str, Any]], metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    file_results: list[FileParseResult] = []
    warnings: list[str] = []
    excluded_parties: dict[str, str] = {}
    meta = metadata or {}
    _collect_excluded_parties_from_source(meta, excluded_parties)
    files = _normalize_input_files(files)
    logger.info("[BankReconciliationDetail] received_files=%s", len(files or []))
    if not files:
        logger.error("[BankReconciliationDetail] failed reason=no_files_received metadata_keys=%s", sorted(meta.keys()))
        return _failure_result("未收到可解析的银行对账明细文件")
    for item in files:
        file_path = item.get("file_path") or ""
        filename = item.get("file_name") or Path(file_path).name
        path_obj = Path(file_path)
        logger.info(
            "[BankReconciliationDetail] file_start filename=%s file_path=%s exists=%s suffix=%s",
            filename,
            file_path,
            path_obj.exists() if file_path else False,
            path_obj.suffix.lower() if file_path else "",
        )
        if not file_path:
            reason = f"{filename or '上传文件'} 未提供文件路径"
            logger.error("[BankReconciliationDetail] failed reason=file_path_missing filename=%s", filename)
            warnings.append(reason)
            continue
        if not path_obj.exists():
            reason = f"文件路径不存在：{file_path}"
            logger.error("[BankReconciliationDetail] failed reason=file_path_not_exists filename=%s path=%s", filename, file_path)
            warnings.append(reason)
            continue
        _collect_excluded_parties_from_source(item, excluded_parties)
        try:
            if path_obj.suffix.lower() == ".pdf":
                pdf_meta = {**meta, **item}
                parsed_pdf, pdf_warnings = parse_pdf_bank_reconciliation_detail(file_path, filename, pdf_meta)
                warnings.extend(pdf_warnings)
                if parsed_pdf:
                    file_results.append(parsed_pdf)
                    logger.info(
                        "[BankReconciliationDetail] pdf_file_parsed file=%s bank=%s page_sheet=%s detail_rows=%s amount_status=%s",
                        filename,
                        parsed_pdf.bank_name,
                        parsed_pdf.sheet_name,
                        len(parsed_pdf.transactions),
                        "完整" if all(tx.get("amount") is not None for tx in parsed_pdf.transactions) else "部分缺失",
                    )
                continue
            sheets = _read_workbook(file_path, filename)
            logger.info(
                "[BankReconciliationDetail] workbook file=%s path=%s sheets=%s",
                filename,
                file_path,
                [(sheet_name, len(rows), max((len(row) for row in rows), default=0)) for sheet_name, rows in sheets],
            )
            matched_file = False
            for sheet_name, rows in sheets:
                parsed = _parse_sheet(rows, sheet_name, filename)
                if parsed:
                    matched_file = True
                    file_results.append(parsed)
                    logger.info(
                        "[BankReconciliationDetail] file=%s sheet=%s bank=%s header_row=%s header_col_start=%s detail_rows=%s amount_status=%s placeholders=%s",
                        filename,
                        sheet_name,
                        parsed.bank_name,
                        parsed.header_row_no,
                        parsed.header_col_start,
                        len(parsed.transactions),
                        "完整" if all(tx.get("amount") is not None for tx in parsed.transactions) else "部分缺失",
                        parsed.placeholder_cleaned_count,
                    )
            if not matched_file:
                logger.warning("[BankReconciliationDetail] file_no_parsed_sheet file=%s path=%s", filename, file_path)
                warnings.append(f"{filename} 表头扫描失败，已扫描工作表前 50 行 50 列，未找到银行对账明细交易表头")
        except Exception as exc:
            logger.exception("[BankReconciliationDetail] parse_failed file=%s", filename)
            warnings.append(f"{filename} 解析失败：{exc}")
    if not file_results:
        reason = warnings[0] if warnings else "未找到交易表头或未读取到有效交易明细"
        logger.error("[BankReconciliationDetail] failed parsed_files=0 reason=%s warnings=%s", reason, warnings)
        data = _failure_result(reason, "请检查文件是否为银行账户明细、交易流水、单位活期账户交易明细或银行对账明细。")
        data["warnings"] = list(dict.fromkeys([*data.get("warnings", []), *warnings]))
        return _json_safe(data)

    data = _aggregate(file_results, excluded_parties=excluded_parties)
    data["warnings"] = list(dict.fromkeys([*data.get("warnings", []), *warnings]))
    data["display_markdown"] = _render_compact_display_markdown(data)
    data["markdown"] = data["display_markdown"]
    data["markdown_summary"] = data["display_markdown"]
    logger.info(
        "[BankReconciliationDetail] aggregate raw=%s deduped=%s duplicate=%s",
        (data.get("summary") or {}).get("raw_transaction_count"),
        (data.get("summary") or {}).get("deduped_transaction_count"),
        (data.get("summary") or {}).get("duplicate_transaction_count"),
    )
    return _json_safe(data)


class BankReconciliationDetailSkill(BaseExtractionSkill):
    document_type = DOC_TYPE
    supported_extensions = {".xlsx", ".xls", ".csv", ".pdf"}
    skill_name = SKILL_NAME
    skill_version = "v1"

    def extract(self, input_data: ExtractionInput) -> ExtractionResult:
        files = _normalize_input_files(input_data.metadata.get("files")) if isinstance(input_data.metadata.get("files"), list) else []
        if not files:
            if input_data.file_path:
                files = [{"file_path": input_data.file_path, "file_name": input_data.file_name}]
        logger.info(
            "[BankReconciliationDetailSkill] start received_file_count=%s received_file_paths=%s source_file_names=%s",
            len(files),
            [item.get("file_path") for item in files],
            [item.get("file_name") for item in files],
        )
        data = parse_bank_reconciliation_files(files, metadata=input_data.metadata)
        markdown = data.get("display_markdown") or ""
        success = data.get("extraction_status") != "failed" and bool((data.get("summary") or {}).get("deduped_transaction_count"))
        return ExtractionResult(
            document_type=DOC_TYPE,
            schema_version=SCHEMA_VERSION,
            extracted_json=data,
            markdown_summary=markdown,
            confidence=0.92 if success else 0.2,
            warnings=list(data.get("warnings") or []),
            errors=[] if success else [str(data.get("failure_reason") or "未识别到有效银行对账明细")],
            skill_name=SKILL_NAME,
            skill_version=self.skill_version,
        )


def build_bank_reconciliation_detail_content(
    *,
    file_path: str,
    file_name: str = "",
    customer_id: str = "",
    document_id: str = "",
    metadata: dict[str, Any] | None = None,
) -> dict[str, Any]:
    skill = BankReconciliationDetailSkill()
    result = skill.extract(
        ExtractionInput(
            customer_id=customer_id,
            document_id=document_id,
            document_type=DOC_TYPE,
            file_name=file_name or Path(file_path).name,
            file_path=file_path,
            mime_type=None,
            raw_text="",
            metadata=metadata or {},
        )
    )
    content = {
        "type": DOC_TYPE,
        "name": DOC_TYPE_NAME,
        "title": DOC_TYPE_NAME,
        "doc_type": DOC_TYPE,
        "doc_type_name": DOC_TYPE_NAME,
        "document_type": DOC_TYPE,
        "document_type_code": DOC_TYPE,
        "document_type_name": DOC_TYPE_NAME,
        "storage_label": DOC_TYPE_NAME,
        "extraction_status": "success" if not result.errors else "failed",
        "extraction_error": "；".join(result.errors),
        "display_markdown": result.markdown_summary,
        "markdown": result.markdown_summary,
        "markdown_summary": result.markdown_summary,
        "report_markdown": result.markdown_summary,
        "structured_data": result.extracted_json,
        "data": {"display_markdown": result.markdown_summary},
    }
    return content
