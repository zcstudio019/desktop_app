from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from backend.services.enterprise_bank_statement_agent import run_enterprise_bank_statement_agent
from backend.services.enterprise_bank_statement_agent.normalizer import normalize_account_number, normalize_amount, normalize_text


AMOUNT_TOLERANCE = 0.01
EXPECTED_COUNTERPARTY_BANKS = {"中国建设银行", "中国农业银行", "交通银行", "招商银行"}


def _compact(value: Any) -> str:
    return normalize_text(value).replace(" ", "").replace("\u3000", "")


def _clean_label(value: Any) -> str:
    text = _compact(value)
    text = text.rstrip(":：")
    text = text.replace("（", "(").replace("）", ")")
    for token in ("(¥)", "(￥)", "(元)", "(人民币)", "¥", "￥", "元"):
        text = text.replace(token, "")
    return text


def _as_float(value: Any) -> float:
    amount = normalize_amount(value)
    return float(amount or 0)


def _as_int(value: Any) -> int:
    amount = normalize_amount(value)
    return int(amount or 0)


def _fmt_amount(value: Any) -> str:
    return f"{float(value or 0):.2f}"


def _fmt_cell(value: Any) -> str:
    if value in (None, ""):
        return "-"
    return str(value)


def _infer_bank_from_sheet(sheet_name: str) -> str:
    if "民生" in sheet_name:
        return "民生银行"
    if "平安" in sheet_name:
        return "平安银行"
    if "泰隆" in sheet_name:
        return "泰隆银行"
    if "浙江网商" in sheet_name or "网商" in sheet_name:
        return "浙江网商"
    return sheet_name


def _iter_label_value_pairs(rows: list[list[Any]], limit: int = 30) -> list[tuple[str, Any]]:
    pairs: list[tuple[str, Any]] = []
    for row in rows[:limit]:
        cells = list(row)
        for index, cell in enumerate(cells):
            label = _clean_label(cell)
            if not label:
                continue
            value = cells[index + 1] if index + 1 < len(cells) else None
            if value not in (None, ""):
                pairs.append((label, value))
            text = normalize_text(cell)
            for sep in (":", "："):
                if sep in text:
                    left, right = text.split(sep, 1)
                    if left.strip() and right.strip():
                        pairs.append((_clean_label(left), right.strip()))
    return pairs


def _find_first_pair(pairs: list[tuple[str, Any]], labels: set[str]) -> Any:
    for label, value in pairs:
        if label in labels:
            return value
    return None


def _canonical_header(value: Any) -> str | None:
    label = _clean_label(value)
    if not label:
        return None
    if any(key in label for key in ("对方开户行", "对方机构", "对方银行", "收付款方开户行", "对方账户开户行")):
        return "counterparty_bank"
    if any(key in label for key in ("对方账号", "对方账户", "收付款方账号", "交易对手账号")):
        return "counterparty_account"
    if label in {"账号", "本方账号", "企业账号", "账户账号", "银行账号", "账户号码"}:
        return "account_number"
    if label in {"交易日期", "记账日期", "入账日期", "交易时间", "发生日期", "日期", "提交时间"}:
        return "transaction_date"
    if label in {"收入", "收入金额", "贷方金额", "贷方发生额", "借方金额(收)", "借方金额（收）", "入账金额", "发生额收入"}:
        return "inflow"
    if label in {"支出", "支出金额", "借方金额", "借方发生额", "贷方金额(支)", "贷方金额（支）", "出账金额", "发生额支出"}:
        return "outflow"
    if label in {"余额", "账户余额", "交易后余额", "可用余额"}:
        return "balance"
    if label in {"对方户名", "对方名称", "交易对手", "对方账户名称", "收付款方名称"}:
        return "counterparty_name"
    if label in {"摘要", "用途", "交易摘要", "备注", "附言", "交易用途", "交易名称"}:
        return "summary"
    return None


def _find_header_row(rows: list[list[Any]]) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_mapping: dict[int, str] = {}
    best_score = 0
    for row_index, row in enumerate(rows[:40]):
        mapping: dict[int, str] = {}
        for col_index, cell in enumerate(row):
            canonical = _canonical_header(cell)
            if canonical and canonical not in mapping.values():
                mapping[col_index] = canonical
        fields = set(mapping.values())
        score = (
            int("transaction_date" in fields)
            + int("inflow" in fields) * 2
            + int("outflow" in fields) * 2
            + int("balance" in fields)
            + int("account_number" in fields)
        )
        if score > best_score:
            best_index = row_index
            best_mapping = mapping
            best_score = score
        if score >= 4 and ("inflow" in fields or "outflow" in fields):
            return row_index, mapping
    return best_index, best_mapping


def _is_effective_transaction(row: dict[str, Any]) -> bool:
    return _as_float(row.get("inflow")) > 0 or _as_float(row.get("outflow")) > 0


def _sum_detail_rows(rows: list[list[Any]], header_index: int | None, mapping: dict[int, str]) -> dict[str, Any]:
    result = {
        "inflow": 0.0,
        "outflow": 0.0,
        "inflow_count": 0,
        "outflow_count": 0,
        "transaction_count": 0,
        "account_number": None,
    }
    if header_index is None or not mapping:
        return result

    account_values: list[str] = []
    for raw_row in rows[header_index + 1 :]:
        normalized = {field: raw_row[index] if index < len(raw_row) else None for index, field in mapping.items()}
        inflow = _as_float(normalized.get("inflow"))
        outflow = _as_float(normalized.get("outflow"))
        if inflow > 0:
            result["inflow"] += inflow
            result["inflow_count"] += 1
        if outflow > 0:
            result["outflow"] += outflow
            result["outflow_count"] += 1
        if inflow > 0 or outflow > 0:
            result["transaction_count"] += 1
        account = normalize_account_number(normalized.get("account_number"))
        if account:
            account_values.append(account)

    if account_values:
        result["account_number"] = Counter(account_values).most_common(1)[0][0]
    result["inflow"] = round(float(result["inflow"]), 2)
    result["outflow"] = round(float(result["outflow"]), 2)
    return result


def extract_original_sheet_summary(sheet_name: str, rows: list[list[Any]]) -> dict[str, Any]:
    pairs = _iter_label_value_pairs(rows)
    header_index, mapping = _find_header_row(rows)
    detail = _sum_detail_rows(rows, header_index, mapping)
    bank = _infer_bank_from_sheet(sheet_name)

    account_number = None
    if bank in {"民生银行", "浙江网商"}:
        account_number = normalize_account_number(
            _find_first_pair(pairs, {"账号", "企业账号", "本方账号", "账户账号", "银行账号"})
        )
    elif bank == "平安银行":
        account_number = detail.get("account_number")

    if bank == "民生银行":
        original_inflow = _as_float(_find_first_pair(pairs, {"贷方累计发生额"}))
        original_outflow = _as_float(_find_first_pair(pairs, {"借方累计发生额"}))
        inflow_count = _as_int(_find_first_pair(pairs, {"贷方累计笔数"}))
        outflow_count = _as_int(_find_first_pair(pairs, {"借方累计笔数"}))
    elif bank == "泰隆银行":
        original_inflow = _as_float(_find_first_pair(pairs, {"总收入"})) or detail["inflow"]
        original_outflow = _as_float(_find_first_pair(pairs, {"总支出"})) or detail["outflow"]
        inflow_count = _as_int(_find_first_pair(pairs, {"总收入笔数"})) or int(detail["inflow_count"])
        outflow_count = _as_int(_find_first_pair(pairs, {"总支出笔数"})) or int(detail["outflow_count"])
    elif bank == "平安银行":
        original_inflow = detail["inflow"]
        original_outflow = detail["outflow"]
        inflow_count = int(detail["inflow_count"])
        outflow_count = int(detail["outflow_count"])
    elif bank == "浙江网商":
        original_inflow = _as_float(_find_first_pair(pairs, {"借方交易金额"})) or detail["inflow"]
        original_outflow = _as_float(_find_first_pair(pairs, {"贷方交易金额"})) or detail["outflow"]
        inflow_count = _as_int(_find_first_pair(pairs, {"借方交易笔数"})) or int(detail["inflow_count"])
        outflow_count = _as_int(_find_first_pair(pairs, {"贷方交易笔数"})) or int(detail["outflow_count"])
    else:
        original_inflow = detail["inflow"]
        original_outflow = detail["outflow"]
        inflow_count = int(detail["inflow_count"])
        outflow_count = int(detail["outflow_count"])
        account_number = detail.get("account_number")
    original_count = int(detail["transaction_count"]) if bank == "平安银行" else int(inflow_count or 0) + int(outflow_count or 0)

    return {
        "sheet_name": sheet_name,
        "bank_name": bank,
        "account_number": account_number,
        "original_inflow": round(float(original_inflow), 2),
        "original_outflow": round(float(original_outflow), 2),
        "original_inflow_count": inflow_count,
        "original_outflow_count": outflow_count,
        "original_count": original_count,
        "header_index": header_index,
        "column_mapping": mapping,
    }


def extract_original_workbook_summary(excel_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        summaries = []
        for sheet in workbook.worksheets:
            rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
            summaries.append(extract_original_sheet_summary(sheet.title, rows))
        return summaries
    finally:
        workbook.close()


def parse_maybe_json(value: Any) -> Any:
    if not value:
        return None
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return None
    return value


def normalize_extracted_json(raw: Any) -> dict[str, Any]:
    parsed = parse_maybe_json(raw)
    if not isinstance(parsed, dict):
        return {}
    if any(key in parsed for key in ("accounts", "account_statements", "accountStatements")):
        return parsed
    for key in ("extracted_json", "extractedJson", "extracted_data", "extractedData", "data", "result", "payload"):
        nested = parse_maybe_json(parsed.get(key))
        if isinstance(nested, dict) and any(item in nested for item in ("accounts", "account_statements", "accountStatements")):
            return nested
    return parsed


def load_system_extracted_json(excel_path: Path, json_path: Path | None) -> dict[str, Any]:
    if json_path:
        return normalize_extracted_json(json_path.read_text(encoding="utf-8"))
    result = run_enterprise_bank_statement_agent(
        file_path=str(excel_path),
        filename=excel_path.name,
        document_type="enterprise_flow",
        metadata={},
    )
    return normalize_extracted_json(result)


def _first(record: dict[str, Any], keys: tuple[str, ...], default: Any = None) -> Any:
    for key in keys:
        value = record.get(key)
        if value not in (None, ""):
            return value
    return default


def extract_system_accounts(extracted_json: dict[str, Any]) -> list[dict[str, Any]]:
    accounts = (
        extracted_json.get("accounts")
        or extracted_json.get("account_statements")
        or extracted_json.get("accountStatements")
        or []
    )
    normalized = []
    for account in accounts if isinstance(accounts, list) else []:
        if not isinstance(account, dict):
            continue
        inflow_count = _first(account, ("inflow_count", "inflowCount", "income_count", "incomeCount"), 0)
        outflow_count = _first(account, ("outflow_count", "outflowCount", "expense_count", "expenseCount"), 0)
        transaction_count = _first(account, ("transaction_count", "transactionCount"), None)
        normalized.append(
            {
                "raw": account,
                "bank_name": _first(account, ("bank_name", "bankName", "bank"), ""),
                "sheet_name": _first(account, ("sheet_name", "sheetName"), ""),
                "account_number": normalize_account_number(_first(account, ("account_number", "accountNumber"), "")),
                "system_inflow": _as_float(_first(account, ("total_inflow", "totalInflow", "income", "credit_amount_total", "total_credit_amount"), 0)),
                "system_outflow": _as_float(_first(account, ("total_outflow", "totalOutflow", "expense", "debit_amount_total", "total_debit_amount"), 0)),
                "system_count": int(transaction_count if transaction_count not in (None, "") else _as_int(inflow_count) + _as_int(outflow_count)),
                "system_inflow_count": _as_int(inflow_count),
                "system_outflow_count": _as_int(outflow_count),
            }
        )
    return normalized


def _contains_match(left: Any, right: Any) -> bool:
    left_text = str(left or "")
    right_text = str(right or "")
    return bool(left_text and right_text and (left_text in right_text or right_text in left_text))


def match_system_account(original: dict[str, Any], candidates: list[dict[str, Any]], used_indexes: set[int]) -> tuple[int | None, dict[str, Any] | None]:
    sheet_name = original.get("sheet_name")
    bank_name = original.get("bank_name")
    account_number = original.get("account_number")
    for index, account in enumerate(candidates):
        if index not in used_indexes and account.get("sheet_name") == sheet_name:
            return index, account
    for index, account in enumerate(candidates):
        if index in used_indexes:
            continue
        if _contains_match(bank_name, account.get("bank_name")) or _contains_match(sheet_name, account.get("bank_name")):
            return index, account
    for index, account in enumerate(candidates):
        if index in used_indexes:
            continue
        if account_number and normalize_account_number(account.get("account_number")) == normalize_account_number(account_number):
            return index, account
    return None, None


def compare_summaries(original_summaries: list[dict[str, Any]], system_accounts: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    diagnostics: list[str] = []
    used: set[int] = set()
    for original in original_summaries:
        index, system = match_system_account(original, system_accounts, used)
        if index is not None:
            used.add(index)
        if not system:
            rows.append({"original": original, "system": None, "status": "❌ 未匹配"})
            account_list = "\n".join(f"- {item.get('bank_name') or '-'} / {item.get('account_number') or '-'}" for item in system_accounts)
            diagnostics.append(
                f"[DIFF][账户匹配]\nsheet={original['sheet_name']} 未匹配到系统账户\n系统账户列表：\n{account_list}\n可能把对方开户行误识别为本方银行。"
            )
            continue

        inflow_diff = round(original["original_inflow"] - system["system_inflow"], 2)
        outflow_diff = round(original["original_outflow"] - system["system_outflow"], 2)
        count_same = int(original["original_count"]) == int(system["system_count"])
        amount_same = abs(inflow_diff) <= AMOUNT_TOLERANCE and abs(outflow_diff) <= AMOUNT_TOLERANCE
        if amount_same and count_same:
            status = "✅ 一致"
        elif amount_same:
            status = "⚠️ 金额一致，笔数不一致"
        else:
            status = "❌ 不一致"
            diagnostics.append(_build_diff_diagnostic(original, system, inflow_diff, outflow_diff))
        rows.append({"original": original, "system": system, "status": status, "inflow_diff": inflow_diff, "outflow_diff": outflow_diff})
    return rows, diagnostics


def _build_diff_diagnostic(original: dict[str, Any], system: dict[str, Any], inflow_diff: float, outflow_diff: float) -> str:
    sheet = original["sheet_name"]
    lines = [
        f"[DIFF][{sheet}]",
        f"原件收入={_fmt_amount(original['original_inflow'])}",
        f"系统收入={_fmt_amount(system['system_inflow'])}",
        f"收入差额={_fmt_amount(inflow_diff)}",
        f"原件支出={_fmt_amount(original['original_outflow'])}",
        f"系统支出={_fmt_amount(system['system_outflow'])}",
        f"支出差额={_fmt_amount(outflow_diff)}",
        "可能原因：",
        "- 系统账户汇总取值优先级错误，顶部空值覆盖了明细反算值",
        "- 系统没有把收入/支出明细列累计到 total_inflow / total_outflow",
        "- 系统账户匹配到了错误的本方银行账户",
    ]
    if "泰隆" in sheet and abs(float(original["original_outflow"])) > 0 and abs(float(system["system_outflow"])) == 0:
        lines.extend(
            [
                "- 系统用 header_summary.total_outflow=None 覆盖了明细反算支出",
                "- account_summary_skill 字段级 fallback 逻辑错误",
            ]
        )
    return "\n".join(lines)


def render_markdown_table(compare_rows: list[dict[str, Any]]) -> str:
    lines = [
        "| sheet_name | 识别银行 | 账号 | 原件收入 | 系统收入 | 收入差额 | 原件支出 | 系统支出 | 支出差额 | 原件笔数 | 系统笔数 | 是否一致 |",
        "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: | --- |",
    ]
    totals = {
        "original_inflow": 0.0,
        "system_inflow": 0.0,
        "original_outflow": 0.0,
        "system_outflow": 0.0,
        "original_count": 0,
        "system_count": 0,
    }
    for item in compare_rows:
        original = item["original"]
        system = item.get("system") or {}
        inflow_diff = round(original["original_inflow"] - float(system.get("system_inflow") or 0), 2)
        outflow_diff = round(original["original_outflow"] - float(system.get("system_outflow") or 0), 2)
        totals["original_inflow"] += original["original_inflow"]
        totals["system_inflow"] += float(system.get("system_inflow") or 0)
        totals["original_outflow"] += original["original_outflow"]
        totals["system_outflow"] += float(system.get("system_outflow") or 0)
        totals["original_count"] += int(original["original_count"] or 0)
        totals["system_count"] += int(system.get("system_count") or 0)
        lines.append(
            "| {sheet} | {bank} | {account} | {oi} | {si} | {idiff} | {oo} | {so} | {odiff} | {oc} | {sc} | {status} |".format(
                sheet=original["sheet_name"],
                bank=_fmt_cell(system.get("bank_name") or original.get("bank_name")),
                account=_fmt_cell(original.get("account_number") or system.get("account_number")),
                oi=_fmt_amount(original["original_inflow"]),
                si=_fmt_amount(system.get("system_inflow")),
                idiff=_fmt_amount(inflow_diff),
                oo=_fmt_amount(original["original_outflow"]),
                so=_fmt_amount(system.get("system_outflow")),
                odiff=_fmt_amount(outflow_diff),
                oc=original["original_count"],
                sc=int(system.get("system_count") or 0),
                status=item["status"],
            )
        )

    total_inflow_diff = round(totals["original_inflow"] - totals["system_inflow"], 2)
    total_outflow_diff = round(totals["original_outflow"] - totals["system_outflow"], 2)
    amount_same = abs(total_inflow_diff) <= AMOUNT_TOLERANCE and abs(total_outflow_diff) <= AMOUNT_TOLERANCE
    count_same = totals["original_count"] == totals["system_count"]
    total_status = "✅ 一致" if amount_same and count_same else ("⚠️ 金额一致，笔数不一致" if amount_same else "❌ 不一致")
    lines.append(
        "| 合计 | - | - | {oi} | {si} | {idiff} | {oo} | {so} | {odiff} | {oc} | {sc} | {status} |".format(
            oi=_fmt_amount(totals["original_inflow"]),
            si=_fmt_amount(totals["system_inflow"]),
            idiff=_fmt_amount(total_inflow_diff),
            oo=_fmt_amount(totals["original_outflow"]),
            so=_fmt_amount(totals["system_outflow"]),
            odiff=_fmt_amount(total_outflow_diff),
            oc=totals["original_count"],
            sc=totals["system_count"],
            status=total_status,
        )
    )
    return "\n".join(lines)


def render_accounts_summary(system_accounts: list[dict[str, Any]]) -> str:
    lines = [
        "| 识别银行 | sheet_name | 账号 | 系统收入 | 系统支出 | 系统笔数 |",
        "| --- | --- | --- | ---: | ---: | ---: |",
    ]
    for account in system_accounts:
        lines.append(
            f"| {_fmt_cell(account.get('bank_name'))} | {_fmt_cell(account.get('sheet_name'))} | {_fmt_cell(account.get('account_number'))} | "
            f"{_fmt_amount(account.get('system_inflow'))} | {_fmt_amount(account.get('system_outflow'))} | {int(account.get('system_count') or 0)} |"
        )
    leaked = [item for item in system_accounts if item.get("bank_name") in EXPECTED_COUNTERPARTY_BANKS]
    if leaked:
        lines.append("")
        lines.append("> 警告：系统账户列表中出现疑似对方开户行：" + "、".join(str(item.get("bank_name")) for item in leaked))
    return "\n".join(lines)


def save_report(excel_path: Path, table: str, diagnostics: list[str], system_accounts: list[dict[str, Any]]) -> Path:
    output_dir = ROOT / "data" / "debug"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"enterprise_flow_consistency_{timestamp}.md"
    content = [
        "# 企业流水原件一致性校验",
        "",
        f"- 文件名：{excel_path.name}",
        f"- 校验时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 对比结果",
        "",
        table,
        "",
        "## 不一致诊断",
        "",
        "\n\n".join(diagnostics) if diagnostics else "未发现不一致。",
        "",
        "## Agent 输出 accounts 摘要",
        "",
        render_accounts_summary(system_accounts),
        "",
    ]
    output_path.write_text("\n".join(content), encoding="utf-8")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="校验企业流水 Agent 解析结果是否与 Excel 原件汇总一致")
    parser.add_argument("excel_path", help="企业流水 Excel 文件路径")
    parser.add_argument("--json", dest="json_path", help="可选：已保存的 extracted_json JSON 文件路径")
    args = parser.parse_args()

    excel_path = Path(args.excel_path).resolve()
    json_path = Path(args.json_path).resolve() if args.json_path else None
    if not excel_path.exists():
        raise SystemExit(f"Excel 文件不存在：{excel_path}")
    if json_path and not json_path.exists():
        raise SystemExit(f"JSON 文件不存在：{json_path}")

    original_summaries = extract_original_workbook_summary(excel_path)
    extracted_json = load_system_extracted_json(excel_path, json_path)
    system_accounts = extract_system_accounts(extracted_json)
    compare_rows, diagnostics = compare_summaries(original_summaries, system_accounts)
    table = render_markdown_table(compare_rows)
    report_path = save_report(excel_path, table, diagnostics, system_accounts)

    print(table)
    if diagnostics:
        print("\n## 不一致诊断\n")
        print("\n\n".join(diagnostics))
    print(f"\n校验报告已保存：{report_path}")


if __name__ == "__main__":
    main()
